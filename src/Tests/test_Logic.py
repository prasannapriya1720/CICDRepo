<<<<<<< HEAD
import pytest
from src.actions.actions import Actions

# Instantiate singleton
action = Actions()


import os
from openpyxl import load_workbook
from src.Utilities.CommonMethods import commonMethods

def get_test_data_from_excel():
    test_data = []

    rootpath = commonMethods.get_project_root(None)
    datapath = os.path.join(rootpath, "Data", "CTKMS_R1_TestData.xlsx")
    workbook = load_workbook(datapath)

    testBrowsers = workbook["TestBrowsers"]
    testModules = workbook["TestModules"]
    testCases = workbook["TestCases"]

    for browser_row in testBrowsers.iter_rows(min_row=2, values_only=True):
        browser_name, browser_run = browser_row[:2]
        if browser_run != 'Y':
            continue

        for module_row in testModules.iter_rows(min_row=2, values_only=True):
            main_module, modulename, module_description, module_run = module_row[:4]
            if module_run != 'Y':
                continue

            for test_case_row in testCases.iter_rows(min_row=2, values_only=True):
                TestCaseName, TestCaseDescription, Run = test_case_row[:3]
                if modulename not in TestCaseName or Run != 'Y':
                    continue

                testStepSheet = workbook[main_module]
                for step_row in testStepSheet.iter_rows(min_row=2, values_only=True):
                    if TestCaseName not in step_row[0] or step_row[4] != 'Y':
                        continue

                    step_data = {
                        "browser": browser_name,
                        "modulename": modulename,
                        "TestCaseName": TestCaseName,
                        "TestCaseDescription": TestCaseDescription,
                        "TestStepName": step_row[0],
                        "TestStepID": step_row[1],
                        "Requirement": step_row[2],
                        "Description": step_row[3],
                        "Keywords": step_row[5],
                        "Locator": step_row[6],
                        "Testdata": step_row[7],
                        "index": step_row[8],
                        "key": step_row[9]
                    }
                    test_data.append(step_data)

    return test_data



@pytest.mark.parametrize("step", get_test_data_from_excel())
def test_excel_driven_execution(step):
    driver = None
    try:
        keyword = step["Keywords"]

        if keyword == 'browser':
            driver = action.invokeBrowser(step["browser"], step["modulename"], step["TestCaseName"],
                                              step["TestStepName"], step["TestStepID"], step["Requirement"],
                                              step["Description"], keyword, step["Locator"],
                                              step["Testdata"], step["TestCaseDescription"])

        elif keyword == 'enter_unique_text1':
            driver = action.enter_unique_text1(step["browser"], step["modulename"], step["TestCaseName"],
                                     step["TestStepName"], step["TestStepID"], step["Requirement"],
                                     step["Description"], keyword, step["Locator"],
                                     step["Testdata"], step["TestCaseDescription"])

        # TODO: Add more keywords and corresponding Actions methods as needed

    except Exception as e:
        print(f"\n❌ Error in step '{step['TestStepName']}' for browser '{step['browser']}': {str(e)}")
        raise

    finally:
        if driver:
            driver.quit()
=======
import glob

import os.path

import openpyxl
import pytest
# from ipython_genutils.py3compat import xrange
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
import time
import unittest

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement

from src.actions.actions import Actions
from src.Utilities.CommonMethods import commonMethods
from src.Utilities.reports import Reports
from datetime import datetime
from configparser import ConfigParser

"""
from configparser import ConfigParser

# importing from config/config.ini
config = ConfigParser()
# print(config.read('src//config//config.ini'))
print(config.read(os.path.expanduser('src/config/config.ini')))
url = config.get('GlobalComponents', 'url')
username = config.get('GlobalComponents', 'username')
password = config.get('GlobalComponents', 'password')
print(url)
print(username)
print(password)

"""


class LogicTest(unittest.TestCase):
    modulename = ""
    TestCaseName = ""
    TestFail = 0
    driver = None
    teststeps_starttime_str = ""
    action = Actions()

    @classmethod
    def setUpClass(cls):
        print("setupclass")
        """
        cls.driver = webdriver.Chrome()
        # executable_path=r"D:\\Software Testing\\SeleniumPythonFramework\\src\\Drivers\\chromedriver.exe")
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()
    
        
        # importing from config/config.ini
        config = ConfigParser()
        # this read location has to be changed if test_Logic file is moved to different location
        print(config.read(os.path.join(os.path.dirname(__file__), '..', 'config', 'config.ini')))
        HCP_url = config.get('RUN', 'HCP_url')
        username = config.get('RUN', 'username')
        password = config.get('RUN', 'password')
        print(url)
        print(username)
        print(password)
        """

    def test_logic(self):
        try:
            # driver = None
            # driver.implicitly_wait(10)
            # driver.maximize_window()
            overall_startTime = datetime.now()

            rootpath = ""
            datapath = "/Data/CTKMS_R1_TestData.xlsx"
            # datapath = "\\Data\\eTMF_TestData.xlsx"
            # datapath = "\\Data\\V-Assure_TestData.xlsx"
            # datapath = "\\Data\\AGGrid_TestData.xlsx"
            # datapath = "\\Data\\GlobalComponents_TestData.xlsx"
            # datapath = "\\Data\\Cybergrants_TestData.xlsx"
            reports = Reports()

            htmlBodyModule = ""
            htmlBodyBrowser = ""
            BrowserName = "Chrome"  # provide the browser name here for now

            rootpath = commonMethods.get_project_root(self)
            # loading workbook using rootpath + datapath
            workbook = load_workbook(rootpath + datapath)

            testSettings = workbook["TestSettings"]
            projectName = testSettings.cell(2, 2).value
            userRequested = testSettings.cell(3, 2).value
            environment = testSettings.cell(4, 2).value
            release = testSettings.cell(5, 2).value
            systemRun = testSettings.cell(6, 2).value
            screenshotFlag = testSettings.cell(7, 2).value

            # print(str(projectName) + "\n" + str(userRequested) + "\n" + str(environment) + "\n" + str(release) + "\n" +
            # str(systemRun))

            testBrowsers = workbook["TestBrowsers"]
            for browserExcel in testBrowsers.iter_rows(values_only=True):
                browser = browserExcel

                browserNameFromExcel = browser.__getitem__(0)
                browser_run = browser.__getitem__(1)
                if browser_run == 'Y':
                    browser_starttime = datetime.now()
                    browser_starttime_str = browser_starttime.strftime("%d/%m/%Y %H:%M:%S")

                    testmodule = workbook["TestModules"]
                    modules = []
                    for row in testmodule.iter_rows(values_only=True):
                        modules = row
                        # print("Modules -> " + str(modules))
                        # global modulename
                        main_module = modules.__getitem__(0)
                        modulename = modules.__getitem__(1)
                        module_description = modules.__getitem__(2)
                        module_run = modules.__getitem__(3)
                        if module_run == 'Y':
                            module_starttime = datetime.now()
                            module_starttime_str = module_starttime.strftime("%d/%m/%Y %H:%M:%S")

                            testcase = workbook["TestCases"]

                            # Check if there is at least one TestCases with "Y" for the current Test module
                            has_y_test = any(modulename in test_row[0] and test_row[2] == 'Y' for test_row in
                                             testcase.iter_rows(min_row=2, values_only=True))

                            # print(f"y for test cases {has_y_test}")
                            if has_y_test:

                                for r in range(2, testcase.max_row + 1):
                                    # print("Row value: " + str(r) + " It has first cell value: " + str(
                                    # testcase.cell(row=r, column=r - 1).value))
                                    testcaselist = []
                                    for j in range(1, testcase.max_column + 1):
                                        # print("col index: " + str(j))

                                        criteria = testcase.cell(row=r, column=j).value
                                        if criteria is None:
                                            criteria = None
                                        else:
                                            criteria = testcase.cell(row=r, column=j).value
                                        testcaselist.insert(j, criteria)
                                    # print(testcaselist)
                                    global TestCaseName
                                    TestCaseName = testcaselist.__getitem__(0)
                                    TestCaseDescription = testcaselist.__getitem__(1)
                                    Run = testcaselist.__getitem__(2)
                                    # TestCase_Summary = testcaselist.__getitem__(3)
                                    # print(f"TestCasename is {TestCaseName}, TestDescription is {TestCaseDescription}, Rus is {Run}")
                                    if modulename in TestCaseName and Run == 'Y':
                                        testcases_starttime = datetime.now()
                                        testcases_starttime_str = testcases_starttime.strftime("%d/%m/%Y %H:%M:%S")

                                        # teststep = workbook["TestSteps"]
                                        teststep = workbook[main_module]

                                        # Check if there is at least one TestStep with "Y" for the current Test Case
                                        has_y_step = any(
                                            TestCaseName in step_row[0] and step_row[4] == 'Y' for step_row in
                                            teststep.iter_rows(min_row=2, values_only=True))

                                        # print(f"y for test steps {has_y_step}")
                                        if has_y_step:
                                            my_list = []
                                            for row in teststep.iter_rows(values_only=True):
                                                # check if the matching value from sheet1 is in the first cell of the row
                                                if TestCaseName in row:
                                                    # print the entire row
                                                    my_list = row
                                                    # print(my_list)
                                                    TestStepName = my_list.__getitem__(0)
                                                    TestStepID = my_list.__getitem__(1)
                                                    Requirement = my_list.__getitem__(2)
                                                    Description = my_list.__getitem__(3)
                                                    Run = my_list.__getitem__(4)
                                                    Keywords = my_list.__getitem__(5)
                                                    Locator = my_list.__getitem__(6)
                                                    Testdata = my_list.__getitem__(7)
                                                    index = my_list.__getitem__(8)
                                                    key = my_list.__getitem__(9)



                                                    if Run == 'Y':
                                                        teststeps_starttime = datetime.now()
                                                        self.teststeps_starttime_str = teststeps_starttime.strftime(
                                                            "%d/%m/%Y %H:%M:%S")
                                                        self.perform(browserNameFromExcel, modulename, TestCaseName,
                                                                     TestStepName,
                                                                     TestStepID, Requirement,
                                                                     Description,
                                                                     Keywords, Locator, Testdata, index, key,
                                                                     TestCaseDescription)
                                                        # print("Row is read: " + str(r) + "and action is performed")

                                            # print("call test case")

                                            reports.Report_TestCases(browserNameFromExcel, main_module, modulename,
                                                                     TestCaseName,
                                                                     TestCaseDescription, testcases_starttime_str)
                                htmlBodyModule = htmlBodyModule + reports.Report_Module(browserNameFromExcel,
                                                                                        main_module, modulename,
                                                                                        module_description,
                                                                                        module_starttime_str)
                                reports.Report_ModuleSummary(browserNameFromExcel, modulename, htmlBodyModule,
                                                             module_starttime_str)
                                htmlBodyModule = ""
                    reports.Report_Browser(browserNameFromExcel)  # this get the module details should be ran
                    reports.Report_BrowserSummary(browserNameFromExcel,
                                                  browser_starttime_str)  # this get the module details should be ran
                    reports.Report_BrowserSummary_working(browserNameFromExcel, browser_starttime_str)
            reports.HighReport(projectName, userRequested, environment, release, systemRun, overall_startTime)
            # foldername = r"C:\repos\test-automation\test-automation\src\TestReports\TestReports_06_02_2023_20_03_21" \
            #            r"\TestCases "
            # awscall.new_upload_to_s3(foldername, "myseleniumtesting")
            self.driver.close()
            self.driver.quit()
        except Exception as e:
            print(f"Error in test_logic -> \n {str(e)}")

    def perform(self, browserNameFromExcel, modulename, TestCaseName, TestStepName, TestStepID, Requirement,
                Description,
                Keywords,
                Locator, Testdata, index, key, TestCase_Summary):
        try:

            if Keywords == 'browser':
                self.driver = self.action.invokeBrowser(browserNameFromExcel, modulename, TestCaseName, TestStepName,
                                                        TestStepID,
                                                        Requirement, Description, Keywords, Locator,
                                                        Testdata, TestCase_Summary)

            # if Keywords == 'browser':
            #     self.driver = self.action.invokeBrowser_devices(browserNameFromExcel, modulename, TestCaseName, TestStepName,
            #                                             TestStepID,
            #                                             Requirement, Description, Keywords, Locator,
            #                                             Testdata, TestCase_Summary)

            elif Keywords == 'enter_URL':
                self.driver = self.action.enterUrl(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                   TestStepName,
                                                   TestStepID, Requirement, Description, Keywords, Locator,
                                                   Testdata, TestCase_Summary)

            elif Keywords == 'click':
                self.driver = self.action.click(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                TestStepName,
                                                TestStepID, Requirement, Description, Keywords, Locator,
                                                Testdata, TestCase_Summary)

            elif Keywords == 'click_Submit':
                self.driver = self.action.click_Submit(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName,
                                                       TestStepID, Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary)

            elif Keywords == 'typedata':
                # self.driver.find_element_by_name(Locator).send_keys(Testdata)
                self.driver = self.action.typedata(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                   TestStepName,
                                                   TestStepID, Requirement, Description, Keywords, Locator,
                                                   Testdata, TestCase_Summary)

            elif Keywords == 'type_password':
                # self.driver.find_element_by_name(Locator).send_keys(Testdata)
                self.driver = self.action.typepassword(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName,
                                                       TestStepID, Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary)

            elif Keywords == 'wait':
                self.driver = self.action.wait(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                               TestStepName,
                                               TestStepID, Requirement, Description, Keywords, Locator,
                                               Testdata, TestCase_Summary)

            elif Keywords == 'enter_keydown':
                self.driver = self.action.enter_keydown(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName, TestStepID, Requirement, Description, Keywords,
                                                        Locator,
                                                        Testdata, TestCase_Summary)

            elif Keywords == 'jsclick':
                self.driver = self.action.jsclick(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                  TestStepName,
                                                  TestStepID, Requirement, Description, Keywords, Locator,
                                                  Testdata, TestCase_Summary)

            elif Keywords == 'gettext1':
                self.driver = self.action.gettext1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                   TestStepName,
                                                   TestStepID, Requirement, Description, Keywords, Locator,
                                                   Testdata, TestCase_Summary)

            elif Keywords == 'gettext2':
                self.driver = self.action.gettext2(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                   TestStepName,
                                                   TestStepID, Requirement, Description, Keywords, Locator,
                                                   Testdata, TestCase_Summary)

            elif Keywords == 'comparedata':
                self.driver = self.action.compareData(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName, TestStepID,
                                                      Requirement, Description, Keywords, Locator,
                                                      Testdata, TestCase_Summary)

            elif Keywords == 'gettext_toarray1':
                self.driver = self.action.gettext_toArray1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, TestCase_Summary)

            elif Keywords == 'gettext_toarray2':
                self.driver = self.action.gettext_toArray2(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, TestCase_Summary)

            elif Keywords == 'comparearray':
                self.driver = self.action.compareArray(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName, TestStepID,
                                                       Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary)

            elif Keywords == 'verify_text':
                self.driver = self.action.compareArray(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName, TestStepID,
                                                       Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary)

            elif Keywords == 'verifyobj':
                self.driver = self.action.verifyObj(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                    TestStepName, TestStepID,
                                                    Requirement, Description, Keywords, Locator,
                                                    Testdata, TestCase_Summary)
            elif Keywords == 'verifywindow':
                self.driver = self.action.verifyWindow(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName, TestStepID,
                                                       Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary)
            elif Keywords == 'verifysortorder':
                self.driver = self.action.verifySort(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                     TestStepName, TestStepID,
                                                     Requirement, Description, Keywords, Locator,
                                                     Testdata, index, TestCase_Summary)
            elif Keywords == 'verifycurrenwindow':
                self.driver = self.action.verifyCurrentWindow(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName,
                                                              TestStepID, Requirement, Description, Keywords,
                                                              Locator,
                                                              Testdata, TestCase_Summary)
            elif Keywords == 'clearandtype':
                self.driver = self.action.clearandtype(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName, TestStepID,
                                                       Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary)
            elif Keywords == 'verifyelementnotvisible':
                self.driver = self.action.verifyElementNotVisible(self.driver, browserNameFromExcel, modulename,
                                                                  TestCaseName, TestStepName,
                                                                  TestStepID, Requirement, Description, Keywords,
                                                                  Locator,
                                                                  Testdata, TestCase_Summary)

            elif Keywords == 'retreiveandsetdata':
                self.driver = self.action.retreiveAndSetData(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords,
                                                             Locator,
                                                             Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveandvalidate':
                self.driver = self.action.retreiveAndValidate(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName,
                                                              TestStepID, Requirement, Description, Keywords,
                                                              Locator,
                                                              Testdata, key, TestCase_Summary)

            elif Keywords == 'validatelink':
                self.driver = self.action.validateLink(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName, TestStepID,
                                                       Requirement, Description, Keywords,
                                                       Locator,
                                                       Testdata, TestCase_Summary)

            elif Keywords == 'refresh':
                self.driver = self.action.refresh(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                  TestStepName,
                                                  TestStepID,
                                                  Requirement, Description, Keywords,
                                                  Locator,
                                                  Testdata, key, TestCase_Summary)

            elif Keywords == 'enterdropdownvalueandsetdatatomap':
                self.driver = self.action.enterdropdownvalueandsetdatatomap(self.driver, browserNameFromExcel,
                                                                            modulename,
                                                                            TestCaseName,
                                                                            TestStepName, TestStepID, Requirement,
                                                                            Description,
                                                                            Keywords,
                                                                            Locator,
                                                                            Testdata, key, TestCase_Summary)
            elif Keywords == 'cleartypeandsetdata':
                self.driver = self.action.cleartypeandsetdata(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName,
                                                              TestStepID, Requirement, Description, Keywords,
                                                              Locator,
                                                              Testdata, key, TestCase_Summary)

            elif Keywords == 'scrollinternalwindow':
                self.driver = self.action.scrollInternalWindow(self.driver, browserNameFromExcel, modulename,
                                                               TestCaseName,
                                                               TestStepName,
                                                               TestStepID, Requirement, Description, Keywords,
                                                               Locator,
                                                               Testdata, key, TestCase_Summary)

            elif Keywords == 'contact_info_check_existing':
                self.driver = self.action.contactInfoCheckExisting(self.driver, browserNameFromExcel, modulename,
                                                                   TestCaseName, TestStepName,
                                                                   TestStepID, Requirement, Description, Keywords,
                                                                   Locator,
                                                                   Testdata, key, TestCase_Summary)

            elif Keywords == 'check_question_fieldtype':
                self.driver = self.action.checkQuestionandfieldtype(self.driver, browserNameFromExcel, modulename,
                                                                    TestCaseName, TestStepName,
                                                                    TestStepID, Requirement, Description, Keywords,
                                                                    Locator,
                                                                    Testdata, key, TestCase_Summary)

            elif Keywords == 'contact_info_checkbox':
                self.driver = self.action.contact_info_checkbox(self.driver, browserNameFromExcel, modulename,
                                                                TestCaseName,
                                                                TestStepName,
                                                                TestStepID, Requirement, Description, Keywords,
                                                                Locator,
                                                                Testdata, key, TestCase_Summary)

            elif Keywords == 'upload_file':
                self.driver = self.action.upload_file(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName,
                                                      TestStepID, Requirement, Description, Keywords,
                                                      Locator,
                                                      Testdata, key, TestCase_Summary)

            elif Keywords == 'verify_text_equal':
                self.driver = self.action.verify_text_equal(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName,
                                                            TestStepID, Requirement, Description, Keywords,
                                                            Locator,
                                                            Testdata, key, TestCase_Summary)

            elif Keywords == 'scrolltoview_verify_text':
                self.driver = self.action.scrolltoview_verify_text(self.driver, browserNameFromExcel, modulename,
                                                                   TestCaseName, TestStepName,
                                                                   TestStepID, Requirement, Description, Keywords,
                                                                   Locator,
                                                                   Testdata, key, TestCase_Summary)

            elif Keywords == 'scroll_inner_div':
                self.driver = self.action.scroll_inner_div(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'checkQuestion_internal':
                self.driver = self.action.checkQuestion_internal(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName,
                                                                 TestStepID, Requirement, Description, Keywords,
                                                                 Locator,
                                                                 Testdata, key, TestCase_Summary)

            elif Keywords == 'checkQuestion_internal_para':
                self.driver = self.action.checkQuestion_internal_para(self.driver, browserNameFromExcel, modulename,
                                                                      TestCaseName,
                                                                      TestStepName,
                                                                      TestStepID, Requirement, Description, Keywords,
                                                                      Locator,
                                                                      Testdata, key, TestCase_Summary)

            elif Keywords == 'enableandclickcheckbox':
                self.driver = self.action.enableandclickcheckbox(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName,
                                                                 TestStepID, Requirement, Description, Keywords,
                                                                 Locator,
                                                                 Testdata, key, TestCase_Summary)

            elif Keywords == 'acceptalert':
                self.driver = self.action.acceptalert(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName,
                                                      TestStepID, Requirement, Description, Keywords,
                                                      Locator,
                                                      Testdata, key, TestCase_Summary)

            elif Keywords == 'dismissalert':
                self.driver = self.action.dismissalert(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName,
                                                       TestStepID, Requirement, Description, Keywords,
                                                       Locator,
                                                       Testdata, key, TestCase_Summary)

            elif Keywords == 'alerttext':
                self.driver = self.action.alerttext(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                    TestStepName,
                                                    TestStepID, Requirement, Description, Keywords,
                                                    Locator,
                                                    Testdata, key, TestCase_Summary)


            elif Keywords == 'datepicker':
                self.driver = self.action.datepicker(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                     TestStepName,
                                                     TestStepID, Requirement, Description, Keywords,
                                                     Locator,
                                                     Testdata, key, TestCase_Summary)

            elif Keywords == 'searchfetchedrequestnumber':
                self.driver = self.action.searchFetchedRequestNumber(self.driver, browserNameFromExcel, modulename,
                                                                     TestCaseName, TestStepName,
                                                                     TestStepID, Requirement, Description, Keywords,
                                                                     Locator,
                                                                     Testdata, key, TestCase_Summary)
            elif Keywords == 'retreiveandsettext':
                self.driver = self.action.retreiveAndSetText(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords,
                                                             Locator,
                                                             Testdata, key, TestCase_Summary)

            elif Keywords == 'openinnewtab':
                self.driver = self.action.enterUrlInNewTab(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveandvalidateinternalportal':
                self.driver = self.action.retreiveAndValidateInternalPortalData(self.driver, browserNameFromExcel,
                                                                                modulename, TestCaseName,
                                                                                TestStepName,
                                                                                TestStepID, Requirement, Description,
                                                                                Keywords,
                                                                                Locator,
                                                                                Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveandvalidateintegerdata':
                self.driver = self.action.retreiveAndValidateIntegerData(self.driver, browserNameFromExcel, modulename,
                                                                         TestCaseName,
                                                                         TestStepName,
                                                                         TestStepID, Requirement, Description, Keywords,
                                                                         Locator,
                                                                         Testdata, key, TestCase_Summary)

            elif Keywords == 'enterdropdownontable':
                self.action.enterdropdownontable(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                 TestStepName,
                                                 TestStepID, Requirement, Description, Keywords,
                                                 Locator,
                                                 Testdata, key, TestCase_Summary)

            elif Keywords == 'textareaontable':
                self.action.textareaontable(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                            TestStepName,
                                            TestStepID, Requirement, Description, Keywords,
                                            Locator,
                                            Testdata, key, TestCase_Summary)

            elif Keywords == 'enterdropdownontable1':
                self.action.enterdropdownontable1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                  TestStepName,
                                                  TestStepID, Requirement, Description, Keywords,
                                                  Locator,
                                                  Testdata, key, TestCase_Summary)

            elif Keywords == 'dateaontable':
                self.action.dateaontable(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                         TestStepName,
                                         TestStepID, Requirement, Description, Keywords,
                                         Locator,
                                         Testdata, key, TestCase_Summary)
            elif Keywords == 'textareaontable1':
                self.action.textareaontable1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                             TestStepName,
                                             TestStepID, Requirement, Description, Keywords,
                                             Locator,
                                             Testdata, key, TestCase_Summary)
            elif Keywords == 'textareaontable2':
                self.action.textareaontable2(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                             TestStepName,
                                             TestStepID, Requirement, Description, Keywords,
                                             Locator,
                                             Testdata, key, TestCase_Summary)
            elif Keywords == 'dateaontable1':
                self.action.dateaontable1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                          TestStepName,
                                          TestStepID, Requirement, Description, Keywords,
                                          Locator,
                                          Testdata, key, TestCase_Summary)
            elif Keywords == 'enterdropdownontable2':
                self.action.enterdropdownontable2(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                  TestStepName,
                                                  TestStepID, Requirement, Description, Keywords,
                                                  Locator,
                                                  Testdata, key, TestCase_Summary)

            elif Keywords == 'checkQuestionOnTable_internal':
                self.action.checkQuestionOnTable_internal(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName,
                                                          TestStepID, Requirement, Description, Keywords,
                                                          Locator,
                                                          Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveAndValidateOnTableValue':
                self.action.retreiveAndValidateOnTableValue(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName,
                                                            TestStepID, Requirement, Description, Keywords,
                                                            Locator,
                                                            Testdata, key, TestCase_Summary)
            elif Keywords == 'retreiveAndValidateOnTableUnit':
                self.action.retreiveAndValidateOnTableUnit(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveAndSetTextOnTableValue':
                self.action.retreiveAndSetTextOnTableValue(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveAndSetTextOnTableUnits':
                self.action.retreiveAndSetTextOnTableUnits(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'retreiveAndValidateIntegerDataOnTable':
                self.action.retreiveAndValidateIntegerDataOnTable(self.driver, browserNameFromExcel, modulename,
                                                                  TestCaseName,
                                                                  TestStepName,
                                                                  TestStepID, Requirement, Description, Keywords,
                                                                  Locator,
                                                                  Testdata, key, TestCase_Summary)

            elif Keywords == 'search_requestID':
                self.driver = self.action.search_requestID(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'hidden_field':
                self.driver = self.action.hidden_field(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName,
                                                       TestStepID, Requirement, Description, Keywords,
                                                       Locator,
                                                       Testdata, TestCase_Summary)

            elif Keywords == 'get_physician_name':
                self.driver = self.action.get_physician_name(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords,
                                                             Locator,
                                                             Testdata, key, TestCase_Summary)

            elif Keywords == 'log_text':
                self.driver = self.action.log_text(self.driver, Locator)

            elif Keywords == 'get_alert_and_Compare':
                self.driver = self.action.get_alert_and_Compare(self.driver, browserNameFromExcel, modulename,
                                                                TestCaseName,
                                                                TestStepName,
                                                                TestStepID, Requirement, Description, Keywords,
                                                                Locator,
                                                                Testdata, key, TestCase_Summary)

            elif Keywords == 'clear_fields':
                self.driver = self.action.clear_allfields(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName,
                                                          TestStepID, Requirement, Description, Keywords,
                                                          Locator,
                                                          Testdata, key, TestCase_Summary)

            elif Keywords == 'obj_removed':
                self.driver = self.action.obj_removed(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName,
                                                      TestStepID, Requirement, Description, Keywords,
                                                      Locator,
                                                      Testdata, key, TestCase_Summary)

            elif Keywords == 'wait_with_no_screenshot':
                self.driver = self.action.wait_with_no_screenshot(self.driver, browserNameFromExcel, modulename,
                                                                  TestCaseName,
                                                                  TestStepName,
                                                                  TestStepID, Requirement, Description, Keywords,
                                                                  Locator,
                                                                  Testdata, TestCase_Summary)

            elif Keywords == 'jsclick_with_no_screenshot':
                self.driver = self.action.jsclick(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                  TestStepName,
                                                  TestStepID, Requirement, Description, Keywords, Locator,
                                                  Testdata, TestCase_Summary)

            elif Keywords == 'approval1_QC_check':
                self.driver = self.action.approval1_QC_check(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords, Locator,
                                                             Testdata, TestCase_Summary)

            elif Keywords == 'approval2_BasicReq_RegApproval':
                self.driver = self.action.approval2_BasicReq_RegApproval(self.driver, browserNameFromExcel, modulename,
                                                                         TestCaseName, TestStepName,
                                                                         TestStepID, Requirement, Description, Keywords,
                                                                         Locator,
                                                                         Testdata, TestCase_Summary)

            elif Keywords == 'approval3_med_approval':
                self.driver = self.action.approval3_med_approval(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName,
                                                                 TestStepID, Requirement, Description, Keywords,
                                                                 Locator,
                                                                 Testdata, TestCase_Summary)

            elif Keywords == 'approval4_Reg_approval':
                self.driver = self.action.approval4_Reg_approval(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName,
                                                                 TestStepID, Requirement, Description, Keywords,
                                                                 Locator,
                                                                 Testdata, TestCase_Summary)

            elif Keywords == 'hover_over_element':
                self.driver = self.action.hover_over_element(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords, Locator,
                                                             Testdata, TestCase_Summary)
            elif Keywords == 'hover_over_webpart':
                self.driver = self.action.hover_over_webpart(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords, Locator,
                                                             Testdata, TestCase_Summary)

            elif Keywords == 'clearField':
                self.driver = self.action.clearField(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                     TestStepName,
                                                     TestStepID, Requirement, Description, Keywords, Locator,
                                                     Testdata, TestCase_Summary)

            elif Keywords == 'listsetData1':
                self.driver = self.action.retrieveasetData1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName,
                                                            TestStepID, Requirement, Description, Keywords, Locator,
                                                            Testdata, TestCase_Summary, key)

            elif Keywords == 'image_value':
                self.driver = self.action.retrieveimage(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName,
                                                        TestStepID, Requirement, Description, Keywords, Locator,
                                                        Testdata, TestCase_Summary, key)


            elif Keywords == 'retrievesortorder':
                self.driver = self.action.retrievesortorder(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName,
                                                            TestStepID, Requirement, Description, Keywords, Locator,
                                                            Testdata, TestCase_Summary, key)

            elif Keywords == 'listInactivecount':
                self.driver = self.action.listInactivecount(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName,
                                                            TestStepID, Requirement, Description, Keywords, Locator,
                                                            Testdata, TestCase_Summary, key)

            elif Keywords == 'ValidateInactivedata':
                self.driver = self.action.ValidateInactivedata(self.driver, browserNameFromExcel, modulename,
                                                               TestCaseName,
                                                               TestStepName,
                                                               TestStepID, Requirement, Description, Keywords, Locator,
                                                               Testdata, TestCase_Summary, key)

            elif Keywords == 'listItemcount':
                self.driver = self.action.listItemcount(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName,
                                                        TestStepID, Requirement, Description, Keywords, Locator,
                                                        Testdata, TestCase_Summary, key)

            elif Keywords == 'ValidatelistItemcount':
                self.driver = self.action.ValidatelistItemcount(self.driver, browserNameFromExcel, modulename,
                                                                TestCaseName,
                                                                TestStepName,
                                                                TestStepID, Requirement, Description, Keywords, Locator,
                                                                Testdata, TestCase_Summary, key)


            elif Keywords == 'retrieveSlidercount':
                self.driver = self.action.retrieveSlidercount(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName,
                                                              TestStepID, Requirement, Description, Keywords, Locator,
                                                              Testdata, TestCase_Summary, key)


            elif Keywords == 'PauseOnHover':
                self.driver = self.action.PauseOnHover(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName,
                                                       TestStepID, Requirement, Description, Keywords, Locator,
                                                       Testdata, TestCase_Summary, key)




            elif Keywords == 'PauseOnHoverOFF':
                self.driver = self.action.PauseOnHoverOFF(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName,
                                                          TestStepID, Requirement, Description, Keywords, Locator,
                                                          Testdata, TestCase_Summary, key)

            elif Keywords == 'tagoff':
                self.driver = self.action.tagoff(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                 TestStepName,
                                                 TestStepID, Requirement, Description, Keywords, Locator,
                                                 Testdata, TestCase_Summary, key)

            elif Keywords == 'emailoff':
                self.driver = self.action.emailoff(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                   TestStepName,
                                                   TestStepID, Requirement, Description, Keywords, Locator,
                                                   Testdata, TestCase_Summary, key)


            elif Keywords == 'locationoff':
                self.driver = self.action.locationoff(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName,
                                                      TestStepID, Requirement, Description, Keywords, Locator,
                                                      Testdata, TestCase_Summary, key)


            elif Keywords == 'joiningdatoff':
                self.driver = self.action.joiningdatoff(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName,
                                                        TestStepID, Requirement, Description, Keywords, Locator,
                                                        Testdata, TestCase_Summary, key)

            elif Keywords == 'retrievecontactdetails':
                self.driver = self.action.retrievecontactdetails(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName,
                                                                 TestStepID, Requirement, Description, Keywords,
                                                                 Locator,
                                                                 Testdata, TestCase_Summary, key)

            elif Keywords == 'retrieveasetQuotes':
                self.driver = self.action.retrieveasetQuotes(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords, Locator,
                                                             Testdata, TestCase_Summary, key)


            elif Keywords == 'is_element_present':
                self.driver = self.action.is_element_present(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords, Locator,
                                                             Testdata, TestCase_Summary, key)

            elif Keywords == 'KeyContactsInactive':
                self.driver = self.action.KeyContactsInactive(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName,
                                                              TestStepID, Requirement, Description, Keywords, Locator,
                                                              Testdata, TestCase_Summary, key)


            elif Keywords == 'doubleclick':
                self.driver = self.action.doubleclick(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName,
                                                      TestStepID, Requirement, Description, Keywords, Locator,
                                                      Testdata, TestCase_Summary, key)


            elif Keywords == 'aggrid_features':
                self.driver = self.action.aggrid_features(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName,
                                                          TestStepID, Requirement, Description, Keywords, Locator,
                                                          Testdata, TestCase_Summary, key)


            elif Keywords == 'check_sorting':
                self.driver = self.action.check_sorting(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName,
                                                        TestStepID, Requirement, Description, Keywords, Locator,
                                                        Testdata, TestCase_Summary, key)

            elif Keywords == 'retry_click':
                self.driver = self.action.retry_click(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName,
                                                      TestStepID, Requirement, Description, Keywords, Locator,
                                                      Testdata, TestCase_Summary, max_attempts=3)

            elif Keywords == 'validategriddata':
                self.driver = self.action.validategriddata(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords, Locator,
                                                           Testdata, key, TestCase_Summary)



            elif Keywords == 'load_data_within_10_seconds':
                self.driver = self.action.load_data_within_10_seconds(self.driver, browserNameFromExcel, modulename,
                                                                      TestCaseName,
                                                                      TestStepName,
                                                                      TestStepID, Requirement, Description, Keywords,
                                                                      Locator,
                                                                      Testdata, key, TestCase_Summary)

            elif Keywords == 'columnsearchfilter':
                self.driver = self.action.columnsearchfilter(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName,
                                                             TestStepID, Requirement, Description, Keywords, Locator,
                                                             Testdata, key, TestCase_Summary)


            elif Keywords == 'columnsearchfilterOR':
                self.driver = self.action.columnsearchfilterOR(self.driver, browserNameFromExcel, modulename,
                                                               TestCaseName,
                                                               TestStepName,
                                                               TestStepID, Requirement, Description, Keywords, Locator,
                                                               Testdata, key, TestCase_Summary)


            elif Keywords == 'retreiveAndValidatedescription':
                self.driver = self.action.retreiveAndValidatedescription(self.driver, browserNameFromExcel, modulename,
                                                                         TestCaseName,
                                                                         TestStepName,
                                                                         TestStepID, Requirement, Description, Keywords,
                                                                         Locator,
                                                                         Testdata, key, TestCase_Summary)


            elif Keywords == 'validateandSetdescriptiondata':
                self.driver = self.action.validateandSetdescriptiondata(self.driver, browserNameFromExcel, modulename,
                                                                        TestCaseName,
                                                                        TestStepName,
                                                                        TestStepID, Requirement, Description, Keywords,
                                                                        Locator,
                                                                        Testdata, key, TestCase_Summary)


            elif Keywords == 'validategridsearchdata':
                self.driver = self.action.validategridsearchdata(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName,
                                                                 TestStepID, Requirement, Description, Keywords,
                                                                 Locator,
                                                                 Testdata, key, TestCase_Summary)

            elif Keywords == 'search_negative_case':
                self.driver = self.action.search_negative_case(self.driver, browserNameFromExcel, modulename,
                                                               TestCaseName,
                                                               TestStepName,
                                                               TestStepID, Requirement, Description, Keywords, Locator,
                                                               Testdata, key, TestCase_Summary)

            elif Keywords == 'press_enter_key':
                self.driver = self.action.press_enter_key(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName,
                                                          TestStepID, Requirement, Description, Keywords, Locator,
                                                          Testdata, key, TestCase_Summary)


            elif Keywords == 'refresh_griddata':
                self.driver = self.action.refresh_griddata(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName,
                                                           TestStepID, Requirement, Description, Keywords, Locator,
                                                           Testdata, key, TestCase_Summary)


            elif Keywords == 'validateFilterClearAfterRefresh':
                self.driver = self.action.validateFilterClearAfterRefresh(self.driver, browserNameFromExcel, modulename,
                                                                          TestCaseName,
                                                                          TestStepName,
                                                                          TestStepID, Requirement, Description,
                                                                          Keywords, Locator,
                                                                          Testdata, key, TestCase_Summary)

            elif Keywords == 'checkcharacterlimit':
                self.driver = self.action.checkcharacterlimit(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName, TestStepID, Requirement, Description,
                                                              Keywords, Locator,
                                                              Testdata, key, TestCase_Summary)

            elif Keywords == 'check_excel_download':
                self.driver = self.action.check_excel_download(self.driver, browserNameFromExcel, modulename,
                                                               TestCaseName,
                                                               TestStepName, TestStepID, Requirement, Description,
                                                               Keywords, Locator,
                                                               Testdata, key, TestCase_Summary)

            elif Keywords == 'verifyNoObj':
                self.driver = self.action.verifyNoObj(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName, TestStepID, Requirement, Description, Keywords,
                                                      Locator,
                                                      Testdata, key, TestCase_Summary)

            elif Keywords == 'beforegroupcount':
                self.driver = self.action.beforegroupcount(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'aftergrouping':
                self.driver = self.action.aftergrouping(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName, TestStepID, Requirement, Description, Keywords,
                                                        Locator,
                                                        Testdata, key, TestCase_Summary)



            elif Keywords == 'hover_over_elementLocator':
                self.driver = self.action.hover_over_elementLocator(self.driver, browserNameFromExcel, modulename,
                                                                    TestCaseName,
                                                                    TestStepName, TestStepID, Requirement, Description,
                                                                    Keywords, Locator,
                                                                    Testdata, TestCase_Summary)

            elif Keywords == 'selectrecord_list':
                self.driver = self.action.selectrecord_list(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName, TestStepID, Requirement, Description,
                                                            Keywords, Locator,
                                                            Testdata, key, TestCase_Summary)


            elif Keywords == 'clickdeletebtncancel':
                self.driver = self.action.clickdeletebtncancel(self.driver, browserNameFromExcel, modulename,
                                                               TestCaseName,
                                                               TestStepName, TestStepID, Requirement, Description,
                                                               Keywords, Locator,
                                                               Testdata, key, TestCase_Summary)


            elif Keywords == 'clickdeletebtn':
                self.driver = self.action.clickdeletebtn(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, key, TestCase_Summary)
            elif Keywords == 'configurelimit':
                self.driver = self.action.configurelimit(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, key, TestCase_Summary)


            elif Keywords == 'get_test_data_and_set':
                self.driver = self.action.get_test_data_and_set(self.driver, browserNameFromExcel, modulename,
                                                                TestCaseName,
                                                                TestStepName, TestStepID, Requirement, Description,
                                                                Keywords,
                                                                Locator,
                                                                Testdata, key, TestCase_Summary)

            elif Keywords == 'validate_grid_excel_data':
                self.driver = self.action.validate_grid_excel_data(self.driver, browserNameFromExcel, modulename,
                                                                   TestCaseName,
                                                                   TestStepName, TestStepID, Requirement, Description,
                                                                   Keywords,
                                                                   Locator,
                                                                   Testdata, key, TestCase_Summary)

            elif Keywords == 'switch_iframe':
                self.driver = self.action.switch_iframe(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName, TestStepID, Requirement, Description, Keywords,
                                                        Locator,
                                                        Testdata, key, TestCase_Summary)

            elif Keywords == 'simple_click':
                self.driver = self.action.simple_click(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                       TestStepName, TestStepID, Requirement, Description, Keywords,
                                                       Locator,
                                                       Testdata, key, TestCase_Summary)

            elif Keywords == 'grid_delete':
                self.driver = self.action.grid_delete(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName, TestStepID, Requirement, Description, Keywords,
                                                      Locator,
                                                      Testdata, key, TestCase_Summary)


            elif Keywords == 'Ag_datefields':
                self.driver = self.action.Ag_datefields(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName, TestStepID, Requirement, Description, Keywords,
                                                        Locator,
                                                        Testdata, key, TestCase_Summary)
            elif Keywords == 'Ag_datefields1':
                self.driver = self.action.Ag_datefields1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, key, TestCase_Summary)

            elif Keywords == 'validateDateRangeFilter':
                self.driver = self.action.validateDateRangeFilter(self.driver, browserNameFromExcel, modulename,
                                                                  TestCaseName,
                                                                  TestStepName, TestStepID, Requirement, Description,
                                                                  Keywords,
                                                                  Locator,
                                                                  Testdata, key, TestCase_Summary)

            elif Keywords == 'check_saveclipboard':
                self.driver = self.action.check_saveclipboard(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName, TestStepID, Requirement, Description,
                                                              Keywords,
                                                              Locator,
                                                              Testdata, key, TestCase_Summary)

            elif Keywords == 'zoomout_80':
                self.driver = self.action.zoomout_80(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                     TestStepName, TestStepID, Requirement, Description, Keywords,
                                                     Locator,
                                                     Testdata, key, TestCase_Summary)

            elif Keywords == 'AGgrid_waitforpageload':
                self.driver = self.action.AGgrid_waitforpageload(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName, TestStepID, Requirement, Description,
                                                                 Keywords,
                                                                 Locator,
                                                                 Testdata, key, TestCase_Summary)


            elif Keywords == 'Veeva_jsclick':
                self.driver = self.action.Veeva_jsclick(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName, TestStepID, Requirement, Description, Keywords,
                                                        Locator,
                                                        Testdata, key, TestCase_Summary)


            elif Keywords == 'eTMF_upload_file':
                self.driver = self.action.eTMF_upload_file(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, key, TestCase_Summary)

            elif Keywords == 'gettext_and_set':
                self.driver = self.action.gettext_and_set(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName, TestStepID, Requirement, Description, Keywords,
                                                          Locator,
                                                          Testdata, key, TestCase_Summary)


            elif Keywords == 'searchwithretreivedata':
                self.driver = self.action.searchwithretreivedata(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName, TestStepID, Requirement, Description,
                                                                 Keywords,
                                                                 Locator,
                                                                 Testdata, key, TestCase_Summary)


            elif Keywords == 'hover_over_eTMFdoc':
                self.driver = self.action.hover_over_eTMFdoc(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName, TestStepID, Requirement, Description,
                                                             Keywords,
                                                             Locator,
                                                             Testdata, TestCase_Summary)

            elif Keywords == 'hover_over_eTMF_Log':
                self.driver = self.action.hover_over_eTMF_Log(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName, TestStepID, Requirement, Description,
                                                              Keywords,
                                                              Locator,
                                                              Testdata, TestCase_Summary)

            elif Keywords == 'select_todaydate':
                self.driver = self.action.select_todaydate(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, TestCase_Summary)


            elif Keywords == 'move_dialog_up':
                self.driver = self.action.move_dialog_up(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'validateStudyNumber_InGrid':
                self.driver = self.action.validateStudyNumber_InGrid(self.driver, browserNameFromExcel, modulename,
                                                                     TestCaseName,
                                                                     TestStepName, TestStepID, Requirement, Description,
                                                                     Keywords,
                                                                     Locator,
                                                                     Testdata, TestCase_Summary)

            elif Keywords == 'hover_over_eTMF_CreateDraft':
                self.driver = self.action.hover_over_eTMF_CreateDraft(self.driver, browserNameFromExcel, modulename,
                                                                      TestCaseName,
                                                                      TestStepName, TestStepID, Requirement,
                                                                      Description, Keywords,
                                                                      Locator,
                                                                      Testdata, TestCase_Summary)

            elif Keywords == 'hover_over_eTMF':
                self.driver = self.action.hover_over_eTMF(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                          TestStepName, TestStepID, Requirement, Description, Keywords,
                                                          Locator,
                                                          Testdata, TestCase_Summary)
            elif Keywords == 'scroll_page_down':
                self.driver = self.action.scroll_page_down(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, TestCase_Summary)

            elif Keywords == 'hover_over_eTMF_ActionMenu':
                self.driver = self.action.hover_over_eTMF_ActionMenu(self.driver, browserNameFromExcel, modulename,
                                                                     TestCaseName,
                                                                     TestStepName, TestStepID, Requirement, Description,
                                                                     Keywords,
                                                                     Locator,
                                                                     Testdata, TestCase_Summary)

            elif Keywords == 'get_StoreData':
                self.driver = self.action.get_StoreData(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                        TestStepName, TestStepID, Requirement, Description, Keywords,
                                                        Locator,
                                                        Testdata, TestCase_Summary)

            elif Keywords == 'enter_unique_text':
                self.driver = self.action.enter_unique_text(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName, TestStepID, Requirement, Description,
                                                            Keywords,
                                                            Locator,
                                                            Testdata, TestCase_Summary)

            elif Keywords == 'enter_unique_text1':
                self.driver = self.action.enter_unique_text1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName, TestStepID, Requirement, Description,
                                                            Keywords,
                                                            Locator,
                                                            Testdata, TestCase_Summary)
            elif Keywords == 'enter_unique_order_number':
                self.driver = self.action.enter_unique_order_number(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName, TestStepID, Requirement, Description,
                                                             Keywords,
                                                             Locator,
                                                             Testdata, TestCase_Summary)

            elif Keywords == 'enter_unique_order_number1':
                self.driver = self.action.enter_unique_order_number1(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName, TestStepID, Requirement, Description,
                                                             Keywords,
                                                             Locator,
                                                             Testdata, TestCase_Summary)


            elif Keywords == 'set_unique_country_code':
                self.driver = self.action.set_unique_country_code(self.driver, browserNameFromExcel, modulename,
                                                                  TestCaseName,
                                                                  TestStepName, TestStepID, Requirement, Description,
                                                                  Keywords,
                                                                  Locator,
                                                                  Testdata, TestCase_Summary)

            elif Keywords == 'select_random_dropdown_value':
                self.driver = self.action.select_random_dropdown_value(self.driver, browserNameFromExcel, modulename,
                                                                       TestCaseName,
                                                                       TestStepName, TestStepID, Requirement,
                                                                       Description, Keywords,
                                                                       Locator,
                                                                       Testdata, TestCase_Summary)

            elif Keywords == 'wait_for_toast_message':
                self.driver = self.action.wait_for_toast_message(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName, TestStepID, Requirement, Description,
                                                                 Keywords,
                                                                 Locator,
                                                                 Testdata, TestCase_Summary)

            elif Keywords == 'verify_storedData':
                self.driver = self.action.verify_storedData(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                            TestStepName, TestStepID, Requirement, Description,
                                                            Keywords,
                                                            Locator,
                                                            Testdata, TestCase_Summary)

            elif Keywords == 'verify_formatted_Date':
                self.driver = self.action.verify_formatted_Date(self.driver, browserNameFromExcel, modulename,
                                                                TestCaseName,
                                                                TestStepName, TestStepID, Requirement, Description,
                                                                Keywords,
                                                                Locator,
                                                                Testdata, TestCase_Summary)

            elif Keywords == 'jsclick_with_retry':
                self.driver = self.action.jsclick_with_retry(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName, TestStepID, Requirement, Description,
                                                             Keywords,
                                                             Locator,
                                                             Testdata, TestCase_Summary)

            elif Keywords == 'is_field_present':
                self.driver = self.action.is_field_present(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, TestCase_Summary)

            elif Keywords == 'search_and_verify_column':
                self.driver = self.action.search_and_verify_column(self.driver, browserNameFromExcel, modulename,
                                                                   TestCaseName,
                                                                   TestStepName, TestStepID, Requirement, Description,
                                                                   Keywords,
                                                                   Locator,
                                                                   Testdata, TestCase_Summary)

            elif Keywords == 'enter_with_retry':
                self.driver = self.action.enter_with_retry(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                           TestStepName, TestStepID, Requirement, Description, Keywords,
                                                           Locator,
                                                           Testdata, TestCase_Summary)

            elif Keywords == 'get_table_data':
                self.driver = self.action.get_table_data(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'get_all_data_of_column':
                self.driver = self.action.get_all_data_of_column(self.driver, browserNameFromExcel, modulename,
                                                                 TestCaseName,
                                                                 TestStepName, TestStepID, Requirement, Description,
                                                                 Keywords,
                                                                 Locator,
                                                                 Testdata, TestCase_Summary)

            elif Keywords == 'enter':
                self.driver = self.action.enter(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                TestStepName, TestStepID, Requirement, Description, Keywords,
                                                Locator,
                                                Testdata, TestCase_Summary)

            elif Keywords == 'verify_table_header':
                self.driver = self.action.verify_table_header(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName, TestStepID, Requirement, Description,
                                                              Keywords,
                                                              Locator,
                                                              Testdata, TestCase_Summary)

            elif Keywords == 'check_action_column':
                self.driver = self.action.check_action_column(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName, TestStepID, Requirement, Description,
                                                              Keywords,
                                                              Locator,
                                                              Testdata, TestCase_Summary)

            elif Keywords == 'pagination':
                self.driver = self.action.pagination(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                     TestStepName, TestStepID, Requirement, Description, Keywords,
                                                     Locator,
                                                     Testdata, TestCase_Summary)

            elif Keywords == 'get_property_value':
                self.driver = self.action.get_property_value(self.driver, browserNameFromExcel, modulename,
                                                             TestCaseName,
                                                             TestStepName, TestStepID, Requirement, Description,
                                                             Keywords,
                                                             Locator,
                                                             Testdata, TestCase_Summary)

            elif Keywords == 'get_CSS_styles':
                self.driver = self.action.get_CSS_styles(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'search_and_verify_all_column':
                self.driver = self.action.search_and_verify_all_column(self.driver, browserNameFromExcel, modulename,
                                                                       TestCaseName,
                                                                       TestStepName, TestStepID, Requirement,
                                                                       Description, Keywords,
                                                                       Locator,
                                                                       Testdata, TestCase_Summary)

            elif Keywords == 'enter_datetime':
                self.driver = self.action.enter_datetime(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'is_field_present_datetime':
                self.driver = self.action.is_field_present_datetime(self.driver, browserNameFromExcel, modulename,
                                                                    TestCaseName,
                                                                    TestStepName, TestStepID, Requirement, Description,
                                                                    Keywords,
                                                                    Locator,
                                                                    Testdata, TestCase_Summary)

            elif Keywords == 'is_field_not_present_datetime':
                self.driver = self.action.is_field_not_present_datetime(self.driver, browserNameFromExcel, modulename,
                                                                    TestCaseName,
                                                                    TestStepName, TestStepID, Requirement, Description,
                                                                    Keywords,
                                                                    Locator,
                                                                    Testdata, TestCase_Summary)

            elif Keywords == 'category_table':
                self.driver = self.action.category_table(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'category_table_edit':
                self.driver = self.action.category_table_edit(self.driver, browserNameFromExcel, modulename,
                                                              TestCaseName,
                                                              TestStepName, TestStepID, Requirement, Description,
                                                              Keywords,
                                                              Locator,
                                                              Testdata, TestCase_Summary)

            elif Keywords == 'category_table_delete':
                self.driver = self.action.category_table_delete(self.driver, browserNameFromExcel, modulename,
                                                                TestCaseName,
                                                                TestStepName, TestStepID, Requirement, Description,
                                                                Keywords,
                                                                Locator,
                                                                Testdata, TestCase_Summary)

            elif Keywords == 'is_disabled':
                self.driver = self.action.is_disabled(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                      TestStepName, TestStepID, Requirement, Description, Keywords,
                                                      Locator,
                                                      Testdata, TestCase_Summary)

            elif Keywords == 'snapshot_check':
                self.driver = self.action.snapshot_check(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'compare_html_css_snapshots':
                self.driver = self.action.compare_html_css_snapshots(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'compare_fetched_column':
                self.driver = self.action.compare_fetched_column(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'Dropdown_value':
                self.driver = self.action.Dropdown_value(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'count_of_elements':
                self.driver = self.action.count_of_elements(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'select_random_dropdown_and_Adhoc':
                self.driver = self.action.select_random_dropdown_and_Adhoc(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'get_all_category_names':
                self.driver = self.action.get_all_category_names(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'get_all_QuestionGroup_questions':
                self.driver = self.action.get_all_QuestionGroup_questions(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'check_cateQuestionGp_with_questions':
                self.driver = self.action.check_cateQuestionGp_with_questions(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'select_random_dropdown_and_Adhoc_DepRule':
                self.driver = self.action.select_random_dropdown_and_Adhoc_DepRule(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'questionGroup':
                self.driver = self.action.questionGroup(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'get_all_QuestionGroup_names':
                self.driver = self.action.get_all_QuestionGroup_names(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'edit_all_categoryGP':
                self.driver = self.action.edit_all_categoryGP(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'search_and_create_Dropdown':
                self.driver = self.action.search_and_create_Dropdown(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'check_and_create_category':
                self.driver = self.action.check_and_create_category(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'search_and_create_Question':
                self.driver = self.action.search_and_create_Question(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'search_and_create_QuestionGroup':
                self.driver = self.action.search_and_create_QuestionGroup(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'Dropdown_select_text':
                self.driver = self.action.Dropdown_select_text(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'select_dropdown_text_and_Adhoc':
                self.driver = self.action.select_dropdown_text_and_Adhoc(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'search_text_and_verify_column':
                self.driver = self.action.search_text_and_verify_column(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'select_dropdown_text_and_Adhoc_Dep_Rule':
                self.driver = self.action.select_dropdown_text_and_Adhoc_Dep_Rule(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'is_field_present_datetime_dropdown':
                self.driver = self.action.is_field_present_datetime_dropdown(self.driver, browserNameFromExcel,
                                                                                  modulename, TestCaseName,
                                                                                  TestStepName, TestStepID, Requirement,
                                                                                  Description, Keywords,
                                                                                  Locator,
                                                                                  Testdata, TestCase_Summary)

            elif Keywords == 'click_and_verify_text':
                self.driver = self.action.click_and_verify_text(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_footer_position':
                self.driver = self.action.verify_footer_position(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)
            elif Keywords == 'verify_welcometext_header':
                self.driver = self.action.verify_welcometext_header(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_reader_landing_page_access':
                self.driver = self.action.verify_reader_landing_page_access(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)
            elif Keywords == 'verify_continent_color_uniqueness':
                self.driver = self.action.verify_continent_color_uniqueness(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)
            elif Keywords == 'verify_continent_tooltips_all':
                self.driver = self.action.verify_continent_tooltips_all(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'scroll_page_down1':
                self.driver = self.action.scroll_page_down1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_inactive_region_selection':
                self.driver = self.action.verify_inactive_region_selection(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_country_overview_submenus':
                self.driver = self.action.verify_country_overview_submenus(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_text_not_present':
                self.driver = self.action.verify_text_not_present(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_level2_submenus_under_level1':
                self.driver = self.action.verify_level2_submenus_under_level1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_level3_questions_under_level2':
                self.driver = self.action.verify_level3_questions_under_level2(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_quick_links_navigation':
                self.driver = self.action.verify_quick_links_navigation(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'select_qg_titles':
                self.driver = self.action.select_qg_titles(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'select_custom_dropdown_value':
                self.driver = self.action.select_custom_dropdown_value(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'hover_and_click':
                self.driver = self.action.hover_and_click(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'hover_select_country':
                self.driver = self.action.hover_select_country(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'hover_select_country1':
                self.driver = self.action.hover_select_country1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)



            elif Keywords == 'verify_floating_menu_text_displayed':
                self.driver = self.action.verify_floating_menu_text_displayed(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_collapsed_nav_icons_present':
                self.driver = self.action.verify_collapsed_nav_icons_present(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_page_title_matches_submenu':
                self.driver = self.action.verify_page_title_matches_submenu(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_search_textbox_visible':
                self.driver = self.action.verify_search_textbox_visible(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_individual_column_search':
                self.driver = self.action.verify_individual_column_search(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)


            elif Keywords == 'verify_action_icons_present':
                self.driver = self.action.verify_action_icons_present(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_unique_ids_in_cr_tasks':
                self.driver = self.action.verify_unique_ids_in_cr_tasks(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_status_column_in_cr_tasks':
                self.driver = self.action.verify_status_column_in_cr_tasks(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_column_values_in_cr_tasks':
                self.driver = self.action.verify_column_values_in_cr_tasks(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_pagination_if_more_than_ten_rows':
                self.driver = self.action.verify_pagination_if_more_than_ten_rows(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_pagination_controlls':
                self.driver = self.action.verify_pagination_controlls(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_popup_alignment_right':
                self.driver = self.action.verify_popup_alignment_right(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_field_alignment_in_popup':
                self.driver = self.action.verify_field_alignment_in_popup(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_default_selected_field':
                self.driver = self.action.verify_default_selected_field(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_only_created_Personal_reports_displayed':
                self.driver = self.action.verify_only_created_Personal_reports_displayed(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'resize_to_device':
                self.driver = self.action.resize_to_device(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_read_only_fields':
                self.driver = self.action.verify_read_only_fields(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_view_report_details':
                self.driver = self.action.verify_view_report_details(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'capture_country_overview_data':
                self.driver = self.action.capture_country_overview_data(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_country_overview_data_in_grid':
                self.driver = self.action.verify_country_overview_data_in_grid(self.driver, browserNameFromExcel, modulename,
                                                                        TestCaseName,
                                                                        TestStepName, TestStepID, Requirement,
                                                                        Description, Keywords,
                                                                        Locator,
                                                                        Testdata, TestCase_Summary)

            elif Keywords == 'compare_country_data_side_by_side':
                self.driver = self.action.compare_country_data_side_by_side(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'compare_export_with_ui':
                self.driver = self.action.compare_export_with_ui(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_export_contains_modified_date':
                self.driver = self.action.verify_export_contains_modified_date(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'verify_maximize_minimize_icon':
                self.driver = self.action.verify_maximize_minimize_icon(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)

            elif Keywords == 'clicksave_wait_for_toast_message1':
                self.driver = self.action.clicksave_wait_for_toast_message1(self.driver, browserNameFromExcel, modulename, TestCaseName,
                                                         TestStepName, TestStepID, Requirement, Description, Keywords,
                                                         Locator,
                                                         Testdata, TestCase_Summary)




            # time.sleep(0.25)
        except Exception as e:
            print(f"Perform method inside test_logic -> {str(e)}")

    def stest_check(self):
        action = Actions()
        self.driver = action.invokeBrowser("chrome", "modulename", "TestCaseName", "TestStepName", "TestStepID",
                                           "Requirement", "Description", "Keywords", "Locator",
                                           "https://sandbox.cybergrants.com/pls/cybergrants-sb/ao_login.login?x_gm_id=9634&x_proposal_type_id=80980",
                                           "TestCase_Summary")
        action.wait(self.driver, "edge", "modulename", "TestCaseName", "TestStepName", "TestStepID",
                    "Requirement", "Description", "Keywords", "Locator",
                    "5",
                    "TestCase_Summary")

    def stest_snapshotTest(self):
        action = Actions()
        action.takesnapshot("https://kms-dev.gilead.com/", "Homepage_current.png",
                            "C:\\Users\\vkumar19\\OneDrive - Gilead Sciences\\Documents\\CTKMS_Release1\\CTKMS_R1_UI_test-automation\\Snapshot_baseline\\Homepage_baseline.png")

    def stest_capture_page_snapshot(self):
        action = Actions()
        action.capture_all_pages()

    @classmethod
    def teardown_class(cls):
        # cls.driver.close()
        # cls.driver.quit()

        # Replace with your SharePoint site URL, client ID, client secret, and library name
        # site_url = 'https://gileadconnect.sharepoint.com/sites/rasbx'
        # client_id = '81c1a8ce-c3cf-4c4d-956f-1c7a2e49c92a'
        # client_secret = 'kdrhdrQ+zPXo4BMfr3ywQQwcP17RlJtc1tBqGM4ijEI='
        # library_name = 'Shared%20Documents/Automation'

        """
        test_reports_path = os.path.join(os.environ.get("Build.SourcesDirectory", ""), "src", "TestReports")
        report_folders = glob.glob(os.path.join(test_reports_path, "*"))
        report_folders.sort(key=os.path.getmtime, reverse=True)
        latest_folder = report_folders[0] if report_folders else None
        print("LATEST_REPORT:", latest_folder)
        """

        # Path to the folder you want to upload
        # local_folder_path = r'C:\repos\test-automation\src\TestReports\TestReports_20_07_2023_00_00_38'

        # Get the path to the current file (SharePoint_upload.py)
        # current_file_path = os.path.abspath(__file__)

        # Go up one level to reach the 'src' folder
        # src_folder_path = os.path.dirname(os.path.dirname(current_file_path))

        # Append 'TestReports' to the path
        # test_reports_folder_path = os.path.join(src_folder_path, 'TestReports')
        # print("Files in 'TestReports' folder:")
        # Find the latest subfolder within 'TestReports'
        # subfolders = [os.path.join(test_reports_folder_path, folder) for folder in os.listdir(test_reports_folder_path)
        #           if os.path.isdir(os.path.join(test_reports_folder_path, folder))]

        # Sort subfolders based on the timestamp in their name
        # latest_folder_path = max(subfolders, key=os.path.getctime)

        # print("latest_folder_path : ", latest_folder_path)
        # sp = SharePointUploader(site_url, client_id, client_secret, library_name)
        # sp.upload_folder(latest_folder_path)

        print("Test Completed")
>>>>>>> 68acd25 (Full project push with updated workflow and long paths enabled)
