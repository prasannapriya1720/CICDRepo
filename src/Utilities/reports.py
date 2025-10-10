import os
import base64
import time
from datetime import datetime
import datetime
from pathlib import Path
from PIL import Image
from Screenshot import Screenshot
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from src.Utilities.CommonMethods import commonMethods
# from src.Tests.test_Logic import LogicTest
from openpyxl.reader.excel import load_workbook
import io
from selenium.webdriver.common.by import By
from io import BytesIO
from PIL import Image, ImageDraw


# from Screenshot import Screenshot_Clipping

# from src.Tests.test_Logic import LogicTest


class Reports:
    _instance = None  # part of singleton instance part of :  def __new__(cls):

    common_methods = commonMethods()
    # test_logic_obj = LogicTest()

    now = datetime.datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")

    formatted_DDMMMYYYY = now.strftime("%d %B %Y")
    formatted_Time_now = now.strftime("%I:%M:%S %p")

    timezone = now.astimezone().tzname()

    # print("date and time =", dt_string)

    # replace all instances of 'r' (old) with 'e' (new)
    dt_string1 = dt_string.replace("/", "_")
    dt_string2 = dt_string1.replace(" ", "_")
    dt_string3 = dt_string2.replace(":", "_")

    # highReport = str(Path(__file__).parent.parent) + "\\TestReports"
    # TestReports = highReport + "\\TestReports_" + dt_string3
    # # highReport file name
    # highReport_filename = TestReports + "\\TestReport.html"
    #
    # TestModuleFolder = TestReports + "\\TestModules"
    # TestCasesFolder = TestReports + "\\TestCases"
    # TestStepsFolder = TestReports + "\\TestSteps"
    # ScreenshotFolder = TestReports + "\\Screenshots"
    # cssFolder = TestReports + "\\css"
    # aboutFolder = TestReports + "\\About"
    # about_the_project_Folder = TestReports + "\\AboutTheProject"
    # TestBrowsersFolder = TestReports + "\\Browsers"
    #
    # highReportWithGraph = TestReports + "\\TestReport_highReport.html"

    highReport = os.path.join(str(Path(__file__).parent.parent), "TestReports")
    TestReports_time = "TestReports_" + dt_string3
    TestReports = os.path.join(highReport, TestReports_time)
    # highReport file name
    highReport_filename = os.path.join(TestReports, "TestReport.html")

    TestModuleFolder = os.path.join(TestReports, "TestModules")
    TestCasesFolder = os.path.join(TestReports, "TestCases")
    TestStepsFolder = os.path.join(TestReports, "TestSteps")
    ScreenshotFolder = os.path.join(TestReports, "Screenshots")
    cssFolder = os.path.join(TestReports, "css")
    aboutFolder = os.path.join(TestReports, "About")
    about_the_project_Folder = os.path.join(TestReports, "AboutTheProject")
    TestBrowsersFolder = os.path.join(TestReports, "Browsers")

    highReportWithGraph = os.path.join(TestReports, "TestReport_highReport.html")

    testdatafilepath = ""
    testcasefilepath = ""
    testmodulefilepath = ""
    Report_TestCase_Filename = ""
    DetailedReport_filename = ""
    Report_TestModule_Filename = ""
    # Status = ""
    detailedReportFail = 0
    numTestDataFail = 0
    numTestDataPass = 0
    module = ""
    numTestCaseFail = 0
    numTestCasePass = 0
    TotalTestCases = 0
    TCModuleName = ""
    TCTestCaseName = ""

    FinalTestCasePass = 0
    FinalTestCaseFail = 0

    browserfilepath = ""
    Report_Browser_Filename = ""

    numModulePass = 0
    numModuleFail = 0
    browser = ""
    TotalModules = 0

    currentTestCaseFail = False
    failedMessage = ""
    screenShotName = ""
    screenshotLocation = ""
    StatusDetbgcolor = "#BCE954"
    stylesString = "<style>" + "table {border-collapse: collapse;width: 100%;}" + "td, th {height: 2rem;border: 1px " \
                                                                                  "solid #ccc;text-align: center;}" + \
                   "th {background: lightblue;border-color: white;}" + "tr:nth-child(even){background-color: " \
                                                                       "#f2f2f2}" + "body {padding: 1rem;}" + \
                   "</style> "

    homepageStyles = "<style>" + \
                     ".TestDetails {border-collapse: collapse;margin-top: -20px !important;height: 235px;}" + \
                     ".TestDetails td {border: 1px solid #ccc;text-align: left;padding: 7px;}" + \
                     ".TestDetails th {border: 1px solid #ccc;text-align: left;padding: 7px;}" + \
                     ".TestDetails td:first-child {width: 150px;}" + \
                     ".TestDetails th {background: lightblue;border-color: white;}" + \
                     ".TestDetails tr:nth-child(even) {background-color: #f2f2f2}" + \
                     "body {padding: 1rem;}" + \
                     ".DetailedReport {border-collapse: collapse;}" + \
                     ".TopBottom {display: block;clear: both;padding-top: 5px;}" + \
                     ".DetailedReport td, th {width: 4rem;height: 2rem;border: 1px solid #ccc;text-align: center;}" + \
                     ".DetailedReport th {background: lightblue;border-color: white;}" + \
                     ".DetailedReport tr:nth-child(even) {background-color: #f2f2f2}" + \
                     "body {padding: 1rem;}" + \
                     "h3.heading {padding: 0px;margin: 0px;margin-bottom: 15px;}" + \
                     ".topLeft {width: 460px;float: left;}" + \
                     ".topRight {margin-left: 480px;}" + \
                     ".Passdiv {background-color: #006633;width: 30px;vertical-align: middle;text-align: " \
                     "center;display: inherit;}" + \
                     ".Faildiv {background-color: #FF0000;width: 30px;vertical-align: middle;text-align: " \
                     "center;display: inherit;}" + \
                     ".ResultSummary {border-collapse: collapse;width: 100%;margin-top: -36px !important;border: 1px " \
                     "solid #d0d0d0;}" + \
                     ".td02 {padding-left: 80px;float: left;height: 195px;}" + \
                     ".t08 {padding-left: 20px;float: left;vertical-align: bottom;height: 195px;}" + \
                     ".PercentageBarFont {color: white;font-weight: bold;}" + \
                     ".t06 {float: left;background-color: antiquewhite;padding: 5px;margin-left: 10px;}" + \
                     "</style>"

    html_body_string = ""
    browserfound = ""
    TCbrowserfound = ""
    TMbrowser = ""
    finalModulePass = 0
    finalModuleFail = 0

    new_html_styles_part1 = "<html><head> " + \
                            "<title>Selenium Python Automation</title>" + \
                            "<link rel=\"stylesheet\" href=\"../css/style.css\">" + \
                            "</head>" + \
                            "<body>" + \
                            "<section id=\"sidebar\">" + \
                            "<div class=\"white-label\">" + \
                            "</div>" + \
                            "<div class=\"container\" id=\"sidebar-nav\">"
    new_html_styles_part2_li = ""


    # Create an empty dictionary for test cases
    # TestCasePass, TestCaseFail, TotalTestCase, SuccessRate, PassWidth, FailRate, FailWidth
    test_cases_map = {}

    main_module_name = ""
    main_module_TC_pass = 0
    main_module_TC_fail = 0

    test_modules_map = {}
    mainModule_name_subModule = ""
    mainModule_subModule_pass = 0
    mainModule_subModule_fail = 0
    report_modulename = ""

    # ____ Screenshot settings______________________
    #datapath = "/Data/eTMF_TestData.xlsx"
    #datapath = "\\Data\\eTMF_TestData.xlsx"
    datapath = os.path.join("Data", "CTKMS_R1_TestData.xlsx")
    rootpath = str(Path(__file__).parent.parent)
    #workbook = load_workbook(rootpath + datapath)
    workbook = load_workbook(os.path.join(rootpath, datapath))

    testSettings = workbook["TestSettings"]
    screenshotFlag = testSettings.cell(7, 2).value
    projectName1 = testSettings.cell(2, 2).value
    # _____________________________________________

    # _______ Product Details _____________________
    productDetails = workbook["CTKMSRelease1"]
    modules_sheet = workbook["TestModules"]
    productname_ProductDetails_sheet = productDetails.cell(2, 1).value
    product_details_columndata = productDetails.cell(2, 2).value
    # _____________________________________________

    new_html_styles_part3 = "</div>" + \
                            "</section>" + \
                            "<section id=\"content-div\">" + \
                            "<div id=\"header\">" + \
                            "<div class=\"header-nav\">" + \
                            "<div class=\"menu-button\">Test Automation Report - " + projectName1 + "</div>" + \
                            "<div class=\"nav\">" + \
                            "<ul>" + \
                            "<li class=\"nav-profile\">" + \
                            "<div class=\"nav-profile-name\">" + \
                            " <i class=\"fa fa-caret-down\"></i>" + \
                            "</div>" + \
                            "</li>" + \
                            "</ul>" + \
                            "</div>" + \
                            "</div>" + \
                            "</div>" + \
                            "<div class=\"content\" id=\"dashboard\" style=\"display: block;\">"

    # Another way is to use a singleton pattern, which will ensure that only one instance of the class is created and
    # shared across all the files.

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __int__(self):
        print()

        # now = datetime.now()
        # dd/mm/YY H:M:S
        # self.dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        # print("date and time =", self.dt_string)

        # highReport = str(commonMethods.get_project_root()) + "\\TestReports"
        # TestReports = self.highReport + "\\TestReports_" + self.dt_string

        # highReport file name
        # highReport_filename = self.TestReports + "\\TestReport.html"

    def createFolder(self, newpath):
        # newpath = commonMethods.get_project_root(self)
        # print(newpath)
        if not os.path.exists(newpath):
            os.makedirs(newpath)

    def folders(self):
        try:
            #print("report folder location outside if: " + self.highReport)
            if not os.path.exists(self.highReport):
                os.makedirs(self.highReport)
                print("report folder location inside if: " + self.highReport)

            if not os.path.exists(self.TestReports):
                os.makedirs(self.TestReports)

            if not os.path.exists(self.TestModuleFolder):
                os.makedirs(self.TestModuleFolder)

            if not os.path.exists(self.TestCasesFolder):
                os.makedirs(self.TestCasesFolder)

            if not os.path.exists(self.TestStepsFolder):
                os.makedirs(self.TestStepsFolder)

            if not os.path.exists(self.ScreenshotFolder):
                os.makedirs(self.ScreenshotFolder)

            if not os.path.exists(self.cssFolder):
                os.makedirs(self.cssFolder)

            if not os.path.exists(self.TestBrowsersFolder):
                os.makedirs(self.TestBrowsersFolder)

            if not os.path.exists(self.aboutFolder):
                os.makedirs(self.aboutFolder)

            if not os.path.exists(self.about_the_project_Folder):
                os.makedirs(self.about_the_project_Folder)
        except Exception as e:
            print(f"'folders' method has this error -> {str(e)}")

    def currentTimeNow(self):
        now = datetime.now()
        # dd/mm/YY H:M:S
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        # print("date and time =", dt_string)

    def CaptureTestImage1(self, driver, browser, testStepName, testStepID, r_element):
        # global screenshotLocation
        screenshot = ""
        self.folders()
        try:
            """
            # get the size of the entire webpage
            width = driver.execute_script("return document.documentElement.scrollWidth")
            height = driver.execute_script("return document.documentElement.scrollHeight")

            # setting the size of the browser window to match the size of the webpage
            driver.set_window_size(width, height)
            """

            ss = Screenshot.Screenshot()
            self.screenShotName = browser + "_" + testStepName + "_" + testStepID + ".png"
            #screenshot = self.ScreenshotFolder + "\\" + self.screenShotName
            screenshot = os.path.join(self.ScreenshotFolder, self.screenShotName)

            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", r_element)
                wait = WebDriverWait(driver, 10)
                wait.until(EC.visibility_of(r_element))
            except Exception as e:
                print()

            ss.full_Screenshot(driver, save_path=self.ScreenshotFolder, image_name=self.screenShotName)

        except Exception as e:
            print("CaptureTestImage Exception Message -> \n" + str(e))
        # return self.screenShotName
        return screenshot

    def CaptureTestImage(self, driver, browser, testStepName, testStepID):
        global screenshotLocation

        self.folders()
        try:
            # Get the height of the entire page
            total_height = int(driver.execute_script("return document.body.scrollHeight"))

            # Set the initial window height
            window_height = int(driver.execute_script("return window.innerHeight"))

            # Create a list to hold the screenshots
            screenshot_list = []

            # Loop through the page and take screenshots
            for i in range(0, total_height, window_height):
                if i + window_height > total_height:
                    # Adjust the window height for the last screenshot
                    window_height = total_height - i
                driver.execute_script("window.scrollTo(0, {});".format(i))
                screenshot = driver.find_element(By.TAG_NAME, 'body').screenshot_as_base64
                screenshot_list.append(screenshot)

            # Merge the screenshots into a single image
            image_list = [Image.open(io.BytesIO(base64.b64decode(screenshot))) for screenshot in
                          screenshot_list]
            image_width, image_height = image_list[0].size
            final_image = Image.new('RGB', (image_width, total_height))
            current_height = 0
            for image in image_list:
                final_image.paste(image, (0, current_height))
                current_height += image.size[1]

            # Save the final image
            self.screenShotName = browser + "_" + testStepName + "_" + testStepID + ".png"
            #screenshotLocation = self.ScreenshotFolder + "\\" + self.screenShotName
            screenshotLocation = os.path.join(self.ScreenshotFolder, self.screenShotName)

            final_image.save(screenshotLocation)
        except Exception as e:
            print("CaptureTestImage Exception Message -> \n" + str(e))
        return screenshotLocation

    def screenshotOfElement1(self, driver, browser, testStepName, testStepID, element):
        # global screenshotLocation

        self.folders()
        try:
            # Scroll the page until the element is in the viewport
            if not element.is_displayed():
                driver.execute_script("arguments[0].scrollIntoView();", element)

            # Focus on the element
            driver.execute_script("arguments[0].focus();", element)

            # Wait for the element to become visible
            wait = WebDriverWait(driver, 30)
            wait.until(EC.visibility_of(element))

            # Get the location and dimensions of the element
            location = element.location
            size = element.size

            # x = location['x'] - 100
            # y = location['y'] - 100
            # width = size['width'] + 200
            # height = size['height'] + 200

            time.sleep(2)
            # Take a screenshot of the element
            screenshot = element.screenshot_as_png
            img = Image.open(BytesIO(screenshot))
            # element_img = img.crop((x, y, x + width, y + height))

            """
            # Annotate the field with a circle
            draw = ImageDraw.Draw(element_img)
            draw.ellipse((0, 0, width, height), outline='green', width=4)
            """

            element_img = img.crop((location['x'], location['y'], location['x'] + size['width'], location['y'] + size['height']))

            # Save the final image
            self.screenShotName = browser + "_" + testStepName + "_" + testStepID + ".png"
            # self.screenshotLocation = self.ScreenshotFolder + "\\" + self.screenShotName
            self.screenshotLocation = os.path.join(self.ScreenshotFolder, self.screenShotName)

            # Save the annotated screenshot
            element_img.save(self.screenshotLocation)

        except Exception as e:
            print("screenshotOfElement Exception Message -> \n" + str(e))
        return self.screenshotLocation

    def screenshotOfElement(self, driver, browser, testStepName, testStepID):
        try:
            # width = driver.execute_script("return Math.max(document.documentElement.clientWidth, window.innerWidth || 0);")
            # height = driver.execute_script("return Math.max(document.documentElement.clientHeight, window.innerHeight || 0);")
            # driver.set_window_size(width, height)

            # Get the location and dimensions of the element
            # location = element.location
            # size = element.size

            # Take a screenshot of the visible area of the page
            screenshot = driver.get_screenshot_as_png()
            img = Image.open(BytesIO(screenshot))

            # draw = ImageDraw.Draw(img)
            # draw.ellipse((0, 0, size['width'], size['height']), outline='green', width=4)

            # Save the final image
            self.screenShotName = browser + "_" + testStepName + "_" + testStepID + ".png"
            # print(f"screenshot name is : {self.screenShotName}")
            # self.screenshotLocation = self.ScreenshotFolder + "\\" + self.screenShotName
            self.screenshotLocation = os.path.join(self.ScreenshotFolder, self.screenShotName)

            # Save the screenshot
            img.save(self.screenshotLocation)
        except Exception as e:
            print("screenshotOfElement Exception Message -> \n" + str(e))
        return self.screenShotName

    def Report_TestDataStep(self, driver, browser, modulename, TestCaseName, TestStepName, TestStepID, Requirement, testStepDesc,
                            Keywords,
                            Locator, ExpectedResult, ActualResult, FlagTestCase, TestCase_Summary):
        # try:
        global DetailedReport_filename
        # global detailedReportFail
        global currentTestCaseFail

        try:
            # print(f"screenshot flag is {screenshotFlag}")
            TSStatus = ""
            browser_without_space = browser.replace(" ", "_")

            self.folders()
            self.testdatafilepath = str(browser_without_space) + "_" + str(modulename) + "_" + str(
                TestCaseName) + "_TestSteps" + ".html"
            #DetailedReport_filename = str(self.TestStepsFolder) + "\\" + str(self.testdatafilepath)
            DetailedReport_filename = os.path.join(str(self.TestStepsFolder), str(self.testdatafilepath))
            Status = FlagTestCase.lower()
            # screenshotlocation = ""

            # reports1 = Reports()

            if self.TCTestCaseName != TestCaseName or self.TCTestCaseName == "":
                self.detailedReportFail = 0
                self.TCTestCaseName = TestCaseName

            if Status == "fail":
                TSStatus = "Fail"
                self.detailedReportFail += 1
                # print("Report_TestDataStep - detailedReportFail inside if -> " + TestStepName + " - " + testStepDesc +
                # " - " + str(self.detailedReportFail))
                currentTestCaseFail = True
                self.failedMessage = testStepDesc
                # self.screenshotlocation = self.CaptureTestImage(driver, browser, TestStepName, TestStepID, r_element)
                # print("Returned screenshot name -> " + screenshotlocation)
            else:
                TSStatus = "Pass"

            if Keywords == "click_Submit" or Keywords == "wait_with_no_screenshot" or Keywords == "jsclick_with_no_screenshot":
               print()
            else:
                if Status == "fail" and self.screenshotFlag == "Fail":
                    self.screenshotlocation = self.screenshotOfElement(driver, browser_without_space, TestStepName, TestStepID)
                elif Status == "pass" and self.screenshotFlag == "Pass":
                    self.screenshotlocation = self.screenshotOfElement(driver, browser_without_space, TestStepName, TestStepID)
                elif self.screenshotFlag == "All":
                    self.screenshotlocation = self.screenshotOfElement(driver, browser_without_space, TestStepName, TestStepID)

            self.StatusDetbgcolor = self.Status_Color(Status)

            # print("if condition for file -> " + str(os.path.exists(self.DetailedReport_filename)))
            if not os.path.exists(DetailedReport_filename):

                # make a new file if not
                tw = open(DetailedReport_filename, 'w', encoding="utf-8")

                tw.write(self.new_html_styles_part1)

                tw.write("<ul>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestReport_highReport.html\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")
                tw.write(
                    "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"test-steps\">Test Steps</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\AboutTheProject\\AboutTheProject.html\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\About\\About.html\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                tw.write("</ul>")
                tw.write(self.new_html_styles_part3)

                tw.write("<h4 align=\"center\"><FONT COLOR=\"660066\" FACE=\"Arial"
                         "\"SIZE=5><b>Details of the Test Case</b></h4>")
                tw.write("<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Module Name:   "
                         + modulename + '<br>' + "Test Case Name:   " + TestCaseName + "</h4> ")

                tw.write("<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Summary of the test case:   "
                         + TestCase_Summary + "</h4> ")

                tw.write("<table class=\"TestDetails\" cellspacing=1 cellpadding=1 border=1 width=100%> <tr>")

                tw.write("<td width=7%  align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Test Step ID</b></td>")
                tw.write("<td class=\"hidden\" width=15% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Keyword</b></td>")
                tw.write("<td width=20% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Requirement</b></td>")
                tw.write("<td width=20% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Action/ Description</b></td>")
                tw.write("<td width=20% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Expected Result</b></td>")
                tw.write("<td width=20% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Actual Result</b></td>")

                # tw.write("<td width=10%  align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                #         "SIZE=2><b>Execution Time</b></td>")

                tw.write("<td width=5% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Status</b></td>")
                tw.write("<td width=8% align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                         "SIZE=2><b>Evidence</b></td></tr>")

                if Status == "fail":
                    tw.write(
                        "<tr><td width=7% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + TestStepID + "</b></td>")
                    tw.write(
                        "<td class=\"hidden\" width=15% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + Keywords +
                        "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + Requirement +
                        "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + testStepDesc +
                        "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + ExpectedResult +
                        "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + ActualResult +
                        "</b></td> ")

                    """
                    teststeps_endTime = datetime.datetime.now()
                    teststeps_endTime_str = teststeps_endTime.strftime("%d/%m/%Y %H:%M:%S")
    
                    teststeps_start_datetime = datetime.datetime.strptime(teststeps_starttime_str, "%d/%m/%Y %H:%M:%S")
                    teststeps_end_datetime = datetime.datetime.strptime(teststeps_endTime_str, "%d/%m/%Y %H:%M:%S")
    
                    tw.write(
                        "<td width=10% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + str(teststeps_end_datetime - teststeps_start_datetime) +
                        "</b></td> ")
                    """
                    tw.write(
                        "<td width=5% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT COLOR=\"#ffffff\" "
                                                                                            "FACE=\"Arial\" "
                                                                                            "SIZE=2><b>" +
                        TSStatus + "</b></td>")
                    # print("screenshotlocation While writing file -> " + self.screenshotlocation)

                    # if Keywords == "snapshot_check":
                    #     tw.write(
                    #         "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                    #         + str(self.screenshotlocation) + ">Baseline-Screenshot</a> <a href = "
                    #         + str(Locator) + ">Current-Screenshot</a> </b></td></tr>")
                    if self.screenshotFlag == "Fail" or self.screenshotFlag == "All":
                        tw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                            + str(self.screenshotlocation) + ">Screenshot</a></b></td></tr>")
                    else:
                        tw.write(
                            "<td width=10% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b></b></td></tr>")

                else:
                    tw.write(
                        "<tr><td width=7% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + TestStepID + "</b></td>")
                    tw.write(
                        "<td class=\"hidden\" width=15% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + Keywords + "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + Requirement + "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + testStepDesc + "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + ExpectedResult + "</b></td> ")
                    tw.write(
                        "<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                        + ActualResult + "</b></td> ")

                    """
                    teststeps_endTime = datetime.datetime.now()
                    teststeps_endTime_str = teststeps_endTime.strftime("%d/%m/%Y %H:%M:%S")
    
                    teststeps_start_datetime = datetime.datetime.strptime(teststeps_starttime_str,
                                                                          "%d/%m/%Y %H:%M:%S")
                    teststeps_end_datetime = datetime.datetime.strptime(teststeps_endTime_str, "%d/%m/%Y %H:%M:%S")
    
                    tw.write(
                        "<td width=10% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + str(
                            teststeps_end_datetime - teststeps_start_datetime) +
                        "</b></td> ")
                    """
                    tw.write(
                        "<td width=5% align=\"center\" bgcolor=" + self.StatusDetbgcolor
                        + "><FONT COLOR=\"#ffffff\" FACE=\"Arial\" SIZE=2><b>" + TSStatus + "</b></td>")

                    # if Keywords == "snapshot_check":
                    #     tw.write(
                    #         "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                    #         + str(self.screenshotlocation) + ">Baseline-Screenshot</a>  <a href = "
                    #         + str(Locator) + ">Current-Screenshot</a> </b></td></tr>")
                    if self.screenshotFlag == "Pass" or self.screenshotFlag == "All":
                        tw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                            + str(self.screenshotlocation) + ">Screenshot</a></b></td></tr>")
                    else:
                        tw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b></b></td></tr>")

                tw.close()

            else:
                # append if already exists

                tsw = open(DetailedReport_filename, 'a')

                if Status == "fail":
                    tsw.write("<tr><td width=7% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + TestStepID + "</b></td>")
                    tsw.write("<td class=\"hidden\" width=15% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + Keywords + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + Requirement + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + testStepDesc + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + ExpectedResult + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + ActualResult + "</b></td> ")

                    """
                    teststeps_endTime = datetime.datetime.now()
                    teststeps_endTime_str = teststeps_endTime.strftime("%d/%m/%Y %H:%M:%S")
    
                    teststeps_start_datetime = datetime.datetime.strptime(teststeps_starttime_str,
                                                                          "%d/%m/%Y %H:%M:%S")
                    teststeps_end_datetime = datetime.datetime.strptime(teststeps_endTime_str, "%d/%m/%Y %H:%M:%S")
    
                    tsw.write(
                        "<td width=10% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + str(
                            teststeps_end_datetime - teststeps_start_datetime) +
                        "</b></td> ")
                    """

                    tsw.write("<td width=5% align=\"center\" bgcolor=" + self.StatusDetbgcolor
                              + "><FONT COLOR=\"#ffffff\" FACE=\"Arial\" SIZE=2><b>" + TSStatus + "</b></td>")

                    # if Keywords == "snapshot_check":
                    #     tsw.write(
                    #         "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                    #         + str(self.screenshotlocation) + ">Baseline-Screenshot</a>  <a href = "
                    #         + str(Locator) + ">Current-Screenshot</a> </b></td></tr>")
                    if self.screenshotFlag == "Fail" or self.screenshotFlag == "All":
                        tsw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                            + str(self.screenshotlocation) + ">Screenshot</a></b></td></tr>")
                    else:
                        tsw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b></b></td></tr>")
                else:
                    tsw.write("<tr><td width=7% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + TestStepID + "</b></td>")
                    tsw.write("<td class=\"hidden\" width=15% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + Keywords + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + Requirement + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + testStepDesc + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + ExpectedResult + "</b></td> ")
                    tsw.write("<td width=20% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>"
                              + ActualResult + "</b></td> ")

                    """
                    teststeps_endTime = datetime.datetime.now()
                    teststeps_endTime_str = teststeps_endTime.strftime("%d/%m/%Y %H:%M:%S")
    
                    teststeps_start_datetime = datetime.datetime.strptime(teststeps_starttime_str,
                                                                          "%d/%m/%Y %H:%M:%S")
                    teststeps_end_datetime = datetime.datetime.strptime(teststeps_endTime_str, "%d/%m/%Y %H:%M:%S")
    
                    tsw.write(
                        "<td width=10% align=\"left\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b>" + str(
                            teststeps_end_datetime - teststeps_start_datetime) +
                        "</b></td> ")
                    """
                    tsw.write("<td width=5% align=\"center\" bgcolor=" + self.StatusDetbgcolor
                              + "><FONT COLOR=\"#ffffff\" FACE=\"Arial\" SIZE=2><b>" + TSStatus + "</b></td>")

                    # if Keywords == "snapshot_check":
                    #     tsw.write(
                    #         "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Sceenshots\\"
                    #         + str(self.screenshotlocation) + ">Baseline-Screenshot</a>  <a href = "
                    #         + str(Locator) + ">Current-Screenshot</a> </b></td></tr>")
                    if self.screenshotFlag == "Pass" or self.screenshotFlag == "All":
                        tsw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b><a href = ..\\Screenshots\\"
                            + str(self.screenshotlocation) + ">Screenshot</a></b></td></tr>")
                    else:
                        tsw.write(
                            "<td width=8% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b></b></td></tr>")

                tsw.close()
        except Exception as e:
            print("'Report_TestDataStep' Exception Message -> " + str(e))

    def Report_TestCases(self, browser, main_module, modulename, TestCaseName, TestCaseDescription,
                         testcases_starttime_str):
        global Report_TestCase_Filename
        global TotalTestCases

        try:
            rstatus = "Pass"
            self.folders()
            # self.testcasefilepath = browser + "_" + str(modulename) + ".html"
            self.testcasefilepath = "TCModuleName.html"
            #Report_TestCase_Filename = str(self.TestCasesFolder) + "\\" + str(self.testcasefilepath)
            Report_TestCase_Filename = os.path.join(str(self.TestCasesFolder), str(self.testcasefilepath))

            if self.TCModuleName != modulename or self.TCModuleName == "":
                TotalTestCases = 0
                self.numModulePass = 0
                self.numModuleFail = 0
            else:
                TotalTestCases += 1

            # print("Report_TestCases - TotalTestCases -> " + str(TotalTestCases))
            # reports2 = Reports()

            # print(
            #    "detailedReportFail - Report_TestCases -> Module name -> " + modulename + " -> Testcasename -> " +
            #    TestCaseName + " detailedReportFail -> " + str(self.detailedReportFail))

            # print("Report_TestCases - detailedReportFail -> " + str(self.detailedReportFail))

            if self.main_module_name != main_module or self.main_module_name == "":
                self.main_module_TC_fail = 0
                self.main_module_TC_pass = 0
                self.main_module_name = main_module
                print(f"main_module_name : {self.main_module_name} and main_module : {main_module} ")

            if self.detailedReportFail > 0:
                rstatus = "Fail"
                self.StatusDetbgcolor = self.Status_Color("fail")
                # numTestDataFail += 1
                self.numTestCaseFail += 1
                self.FinalTestCaseFail += 1
                self.main_module_TC_fail += 1
                # print(f"main module inside -> if self.detailedReportFail > 0: {main_module}")
            else:
                self.StatusDetbgcolor = self.Status_Color("pass")
                # numTestDataPass += 1
                self.numTestCasePass += 1
                self.FinalTestCasePass += 1
                self.main_module_TC_pass += 1
                # print(f"main module inside -> else self.detailedReportFail > 0: {main_module}")

            self.test_cases_map[main_module] = (
                self.main_module_TC_pass, self.main_module_TC_fail, self.main_module_TC_pass + self.main_module_TC_fail)

            if (self.TCbrowserfound != browser or self.TCbrowserfound == "") and (
                    self.TCModuleName != modulename or self.TCModuleName == ""):
                self.TCModuleName = modulename
                self.TCbrowserfound = browser
                print(f"main module is {main_module}")
                # self.test_cases_map[main_module] = (self.FinalTestCasePass, self.FinalTestCaseFail)

                if not os.path.exists(Report_TestCase_Filename):
                    # make a new file if not
                    rb = open(Report_TestCase_Filename, 'w')
                    # rb.write(self.new_html_styles_part1)

                    rb.write("<html><head> ")
                    rb.write("<title>Selenium Python Automation</title>")
                    rb.write("<link rel=\"stylesheet\" href=\"../css/style.css\">")
                    rb.write("<script type=\"text/javascript\">")

                    # function to get all the div and load into the drop down values
                    rb.write("function updateTestStepOptions() {")
                    rb.write("const testStepDivs = document.querySelectorAll('.testsuites-divs');")
                    rb.write(
                        "const browserDropdown = document.querySelector('select[name=\"test-browser-dropdown\"]');")
                    rb.write("const moduleDropdown = document.querySelector('select[name=\"test-module-dropdown\"]');")
                    rb.write("browserDropdown.innerHTML = '<option value=\"\">All</option>';")
                    rb.write("moduleDropdown.innerHTML = '<option value=\"\">All</option>';")
                    rb.write("for (let i = 0; i < testStepDivs.length; i++) {")
                    rb.write("const div = testStepDivs[i];")
                    rb.write("const browser = div.getAttribute('data-browser');")
                    rb.write("const module = div.getAttribute('data-module');")
                    rb.write("if (!browserDropdown.querySelector(`option[value=\"${browser}\"]`)) {")
                    rb.write("const browserOption = document.createElement('option');")
                    rb.write("browserOption.value = browser;")
                    rb.write("browserOption.text = browser;")
                    rb.write("browserDropdown.appendChild(browserOption);")
                    rb.write("}")
                    rb.write("if (!moduleDropdown.querySelector(`option[value=\"${module}\"]`)) {")
                    rb.write("const moduleOption = document.createElement('option');")
                    rb.write("moduleOption.value = module;")
                    rb.write("moduleOption.text = module;")
                    rb.write("moduleDropdown.appendChild(moduleOption);")
                    rb.write("}")
                    rb.write("}")
                    rb.write("}")
                    # ______________________________________

                    # function to filter values based on the drop down
                    rb.write("function filterDivs() {")
                    rb.write("let noResult = 0;")
                    rb.write("const testStepDivs = document.querySelectorAll('.testsuites-divs');")
                    rb.write(
                        "const browserDropdown = document.querySelector('select[name=\"test-browser-dropdown\"]');")
                    rb.write("const moduleDropdown = document.querySelector('select[name=\"test-module-dropdown\"]');")
                    rb.write("const selectedBrowser = browserDropdown.options[browserDropdown.selectedIndex].value;")
                    rb.write("const selectedModule = moduleDropdown.options[moduleDropdown.selectedIndex].value;")
                    rb.write("for (let i = 0; i < testStepDivs.length; i++) {")
                    rb.write("const div = testStepDivs[i];")
                    rb.write("const browser = div.getAttribute('data-browser');")
                    rb.write("const module = div.getAttribute('data-module');")
                    rb.write("let shouldDisplay = true;")
                    rb.write("if (selectedBrowser && selectedBrowser !== browser && selectedBrowser !== 'All') {")
                    rb.write("shouldDisplay = false;")
                    rb.write("}")
                    rb.write("if (selectedModule && selectedModule !== module && selectedModule !== 'All') {")
                    rb.write("shouldDisplay = false;")
                    rb.write("}")
                    rb.write("if (shouldDisplay) {")
                    rb.write("div.style.display = 'block';")
                    rb.write("noResult++;")
                    rb.write("} else {")
                    rb.write("div.style.display = 'none';")
                    rb.write("}")
                    rb.write("}")
                    rb.write("const searchResultsMessage = document.querySelector('.search-results-message');")
                    rb.write("if (noResult === 0) {")
                    rb.write("searchResultsMessage.style.display = 'block';")
                    rb.write("} else {")
                    rb.write("searchResultsMessage.style.display = 'none';")
                    rb.write("}")
                    rb.write("}")
                    # ________________________________________________

                    rb.write("</script>")
                    rb.write("</head>")
                    rb.write("<body onload=\"updateTestStepOptions()\">")
                    rb.write("<section id=\"sidebar\">")
                    rb.write("<div class=\"white-label\">")
                    rb.write("</div>")
                    rb.write("<div class=\"container\" id=\"sidebar-nav\">")

                    rb.write("<ul>")
                    rb.write(
                        "<li class=\"leftnav\"><a href=\"..\\TestReport_highReport.html\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                    rb.write(
                        "<li class=\"leftnav\"><a href=\"..\\Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                    rb.write(
                        "<li class=\"leftnav\"><a href=\"..\\TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                    rb.write(
                        "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")

                    rb.write(
                        "<li class=\"leftnav\"><a href=\"..\\AboutTheProject\\AboutTheProject.html\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                    rb.write(
                        "<li class=\"leftnav\"><a href=\"..\\About\\About.html\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                    rb.write("</ul>")
                    rb.write(self.new_html_styles_part3)

                    rb.write("<div class=\"content-header\">")
                    rb.write("<h1> Test Cases </h1>")
                    rb.write("<p>Test Automation Report - List of Test Cases with Results</p>")
                    rb.write("</div>")

                    # adding drop downs and buttons
                    rb.write("<br>")
                    rb.write(
                        "<select title=\"Select a browser to display\" class=\"browserfilter\" name=\"test-browser-dropdown\">")
                    rb.write("</select>")
                    rb.write(
                        "<select title=\"Select a module to display\" class=\"testmodulefilter\" name=\"test-module-dropdown\">")
                    rb.write("</select>")
                    rb.write("")
                    rb.write("<button onclick=\"filterDivs()\">Search</button>")
                    rb.write("")
                    rb.write("<div class=\"search-results-message\" style=\"display:none;\">")
                    rb.write("<p> There is no data with the search result</p>")
                    rb.write("</div>")

                    rb.write(
                        "<div data-browser=" + browser + " data-module=" + modulename + " style=\"display:none;\" class=\"testsuites-divs\">")
                    rb.write(
                        "<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Browser Name:   " + browser + "<br> Module Name:   "
                        + modulename + "</h4>")
                    rb.write("<table class=\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100%>")
                    rb.write("<tr>")
                    rb.write(
                        "<td width=30% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Test Case "
                        "Name</b></td>")
                    rb.write(
                        "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Description "
                        "</b></td>")
                    rb.write(
                        "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Execution Time "
                        "</b></td>")

                    rb.write(
                        "<td width=20% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Status</b></td>")
                    rb.write("</tr >")

                    rb.write("<tr>")
                    # print("DetailedReport_filename -> " + DetailedReport_filename)
                    rb.write("<td width=30% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b><a href= ..\\TestSteps\\"
                             + self.testdatafilepath + ">" + TestCaseName + "</a></b></td>")
                    # DetailedReport_filename

                    rb.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                             + TestCaseDescription + "</b></td>")

                    testcases_endTime = datetime.datetime.now()
                    testcases_endTime_str = testcases_endTime.strftime("%d/%m/%Y %H:%M:%S")

                    testcases_start_datetime = datetime.datetime.strptime(testcases_starttime_str, "%d/%m/%Y %H:%M:%S")
                    testcases_end_datetime = datetime.datetime.strptime(testcases_endTime_str, "%d/%m/%Y %H:%M:%S")

                    rb.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                             + str(testcases_end_datetime - testcases_start_datetime) + "</b></td>")

                    rb.write("<td width=20% align=center bgcolor=" + self.StatusDetbgcolor +
                             "><FONT COLOR=#ffffff FACE=Arial SIZE=2><b>" + rstatus + "</b></td>")
                    rb.write("</tr>")
                    rb.close()
                else:
                    rb1 = open(Report_TestCase_Filename, 'a')
                    rb1.write(
                        "</table></div><div data-browser=" + browser + " data-module=" + modulename + " style=\"display:none;\" class=\"testsuites-divs\">")
                    rb1.write(
                        "<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Browser Name     :   " + browser + "<br> Module Name      :   "
                        + modulename + "</h4>")
                    rb1.write("<table class=\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100%>")
                    rb1.write("<tr>")
                    rb1.write(
                        "<td width=30% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Test Case "
                        "Name</b></td>")
                    rb1.write(
                        "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Description "
                        "</b></td>")
                    rb1.write(
                        "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Execution Time "
                        "</b></td>")

                    rb1.write(
                        "<td width=20% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Status</b></td>")
                    rb1.write("</tr >")

                    rb1.write("<tr>")
                    # print("DetailedReport_filename -> " + DetailedReport_filename)
                    rb1.write("<td width=30% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b><a href=..\\TestSteps\\"
                             + self.testdatafilepath + ">" + TestCaseName + "</a></b></td>")
                    # DetailedReport_filename

                    rb1.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                              + TestCaseDescription + "</b></td>")
                    testcases_endTime = datetime.datetime.now()
                    testcases_endTime_str = testcases_endTime.strftime("%d/%m/%Y %H:%M:%S")

                    testcases_start_datetime = datetime.datetime.strptime(testcases_starttime_str, "%d/%m/%Y %H:%M:%S")
                    testcases_end_datetime = datetime.datetime.strptime(testcases_endTime_str, "%d/%m/%Y %H:%M:%S")

                    rb1.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                              + str(testcases_end_datetime - testcases_start_datetime) + "</b></td>")
                    rb1.write("<td width=20% align=center bgcolor=" + self.StatusDetbgcolor +
                              "><FONT COLOR=#ffffff FACE=Arial SIZE=2><b>" + rstatus + "</b></td>")
                    rb1.write("</tr>")
                    rb1.close()
            elif self.TCbrowserfound != browser and self.TCModuleName == modulename:
                self.TCModuleName = modulename
                self.TCbrowserfound = browser
                rb1 = open(Report_TestCase_Filename, 'a')
                rb1.write(
                    "</table></div><div data-browser=" + browser + " data-module=" + modulename + " style=\"display:none;\" class=\"testsuites-divs\">")
                rb1.write(
                    "<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Browser Name     :   " + browser + "<br> Module Name      :   "
                    + modulename + "</h4>")
                rb1.write("<table class=\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100%>")
                rb1.write("<tr>")
                rb1.write(
                    "<td width=30% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Test Case "
                    "Name</b></td>")
                rb1.write(
                    "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Description "
                    "</b></td>")

                rb1.write(
                    "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Execution Time "
                    "</b></td>")

                rb1.write(
                    "<td width=20% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Status</b></td>")
                rb1.write("</tr >")

                rb1.write("<tr>")
                # print("DetailedReport_filename -> " + DetailedReport_filename)
                rb1.write("<td width=30% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b><a href=..\\TestSteps\\"
                             + self.testdatafilepath + ">" + TestCaseName + "</a></b></td>")
                # DetailedReport_filename

                rb1.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                          + TestCaseDescription + "</b></td>")

                testcases_endTime = datetime.datetime.now()
                testcases_endTime_str = testcases_endTime.strftime("%d/%m/%Y %H:%M:%S")

                testcases_start_datetime = datetime.datetime.strptime(testcases_starttime_str, "%d/%m/%Y %H:%M:%S")
                testcases_end_datetime = datetime.datetime.strptime(testcases_endTime_str, "%d/%m/%Y %H:%M:%S")

                rb1.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                          + str(testcases_end_datetime - testcases_start_datetime) + "</b></td>")

                rb1.write("<td width=20% align=center bgcolor=" + self.StatusDetbgcolor +
                          "><FONT COLOR=#ffffff FACE=Arial SIZE=2><b>" + rstatus + "</b></td>")
                rb1.write("</tr>")
                rb1.close()
            elif self.TCbrowserfound == browser and self.TCModuleName != modulename:
                self.TCModuleName = modulename
                self.TCbrowserfound = browser
                rb1 = open(Report_TestCase_Filename, 'a')
                rb1.write(
                    "</table></div><div data-browser=" + browser + " data-module=" + modulename + " style=\"display:none;\" class=\"testsuites-divs\">")
                rb1.write(
                    "<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Browser Name     :   " + browser + "<br> Module Name      :   "
                    + modulename + "</h4>")
                rb1.write("<table class=\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100%>")
                rb1.write("<tr>")
                rb1.write(
                    "<td width=30% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Test Case "
                    "Name</b></td>")
                rb1.write(
                    "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Description "
                    "</b></td>")

                rb1.write(
                    "<td width=25% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Execution Time "
                    "</b></td>")

                rb1.write(
                    "<td width=20% align=center  bgcolor=#0274BD><FONT COLOR=#E0E0E0 FACE= Arial SIZE=2><b>Status</b></td>")
                rb1.write("</tr >")

                rb1.write("<tr>")
                # print("DetailedReport_filename -> " + DetailedReport_filename)
                rb1.write("<td width=30% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b><a href=..\\TestSteps\\"
                             + self.testdatafilepath + ">" + TestCaseName + "</a></b></td>")
                # DetailedReport_filename

                rb1.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                          + TestCaseDescription + "</b></td>")

                testcases_endTime = datetime.datetime.now()
                testcases_endTime_str = testcases_endTime.strftime("%d/%m/%Y %H:%M:%S")

                testcases_start_datetime = datetime.datetime.strptime(testcases_starttime_str, "%d/%m/%Y %H:%M:%S")
                testcases_end_datetime = datetime.datetime.strptime(testcases_endTime_str, "%d/%m/%Y %H:%M:%S")

                rb1.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                          + str(testcases_end_datetime - testcases_start_datetime) + "</b></td>")

                rb1.write("<td width=20% align=center bgcolor=" + self.StatusDetbgcolor +
                          "><FONT COLOR=#ffffff FACE=Arial SIZE=2><b>" + rstatus + "</b></td>")
                rb1.write("</tr>")
                rb1.close()
            else:
                rbw = open(Report_TestCase_Filename, 'a')
                rbw.write("<tr>")
                rbw.write("<td width=30% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b><a href=..\\TestSteps\\"
                             + self.testdatafilepath + ">" + TestCaseName + "</a></b></td>")
                rbw.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                          + TestCaseDescription + "</b></td>")

                testcases_endTime = datetime.datetime.now()
                testcases_endTime_str = testcases_endTime.strftime("%d/%m/%Y %H:%M:%S")

                testcases_start_datetime = datetime.datetime.strptime(testcases_starttime_str, "%d/%m/%Y %H:%M:%S")
                testcases_end_datetime = datetime.datetime.strptime(testcases_endTime_str, "%d/%m/%Y %H:%M:%S")

                rbw.write("<td width=25% align=center><FONT COLOR=#153E7E FACE=Arial SIZE=2><b>"
                          + str(testcases_end_datetime - testcases_start_datetime) + "</b></td>")

                rbw.write("<td width=20% align=center bgcolor=" + self.StatusDetbgcolor +
                          "><FONT COLOR=#ffffff FACE=Arial SIZE=2><b>" + rstatus + "</b></td>")
                rbw.write("</tr>")
                rbw.close()
            self.detailedReportFail = 0
        except Exception as e:
            print(f"'Report_TestCases' has this error -> \n {str(e)}")

    def Report_Module(self, browser, main_module, modulename, module_description, module_starttime_str):
        totalTC = 0
        ModuleStatus = ""

        testbody1 = ""
        testbody2 = ""
        testbody3 = ""
        testbody4 = ""
        testbody5 = ""
        testbody21 = ""
        testbody41 = ""
        try:
            # global html_body_string

            # print("module and modulename -> " + self.module + " - " + modulename)
            # print("Report_Module - numTestCaseFail  -> " + str(self.numTestCaseFail))

            if self.mainModule_name_subModule != main_module or self.mainModule_name_subModule == "":
                self.mainModule_subModule_fail = 0
                self.mainModule_subModule_pass = 0
                self.mainModule_name_subModule = main_module
                # print(f"main_module_name : {self.main_module_name} and main_module : {main_module} ")

            if self.TMbrowser != browser or self.TMbrowser == "":
                self.TMbrowser = browser
                self.numModuleFail = 0
                self.numModulePass = 0

            if self.numTestCaseFail > 0:
                ModuleStatus = "Fail"
                self.numModuleFail += 1
                self.finalModuleFail += 1
                self.mainModule_subModule_fail += 1
                self.StatusDetbgcolor = self.Status_Color("fail")
            else:
                ModuleStatus = "Pass"
                self.StatusDetbgcolor = self.Status_Color("pass")
                self.numModulePass += 1
                self.finalModulePass += 1
                self.mainModule_subModule_pass += 1

            self.test_modules_map[main_module] = (
                self.mainModule_subModule_pass, self.mainModule_subModule_fail,
                self.mainModule_subModule_pass + self.mainModule_subModule_fail)

            # if self.module != modulename or self.module == "":
            #   self.numTestCaseFail = 0
            #   self.numTestCasePass = 0
            #   self.module = modulename

            # print("Report_TestCase_Filename in Report_Module -> " + Report_TestCase_Filename)
            # testbody1 = "<tr><td width=20% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b><a href= " + \
            #           Report_TestCase_Filename + ">" + modulename + "</a></b></td>"

            testbody1 = "<tr><td width=20% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b><a href= ..\\TestCases\\TCModuleName.html >" + \
                         modulename + "</a></b></td>"

            # print("str(self.numTestCasePass+self.numTestCaseFail)-> " + str(self.numTestCasePass + self.numTestCaseFail))
            # print("self.numTestCasePass" + str(self.numTestCasePass))
            # print("self.numTestCaseFail)" + str(self.numTestCaseFail))

            testbody21 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                         + module_description + "</b></td>"

            testbody2 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                        + str(self.numTestCasePass + self.numTestCaseFail) + "</b></td>"

            testbody3 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                        + str(self.numTestCasePass) + "</b></td>"

            testbody4 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                        + str(self.numTestCaseFail) + "</b></td>"

            module_endTime = datetime.datetime.now()
            module_endTime_str = module_endTime.strftime("%d/%m/%Y %H:%M:%S")

            module_start_datetime = datetime.datetime.strptime(module_starttime_str, "%d/%m/%Y %H:%M:%S")
            module_end_datetime = datetime.datetime.strptime(module_endTime_str, "%d/%m/%Y %H:%M:%S")

            testbody41 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                         + str(module_end_datetime - module_start_datetime) + "</b></td>"

            testbody5 = "<td width=15% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT COLOR=\"#FFFFFF\" " \
                                                                                            "FACE=\"Arial\" SIZE=2><b>" + \
                        ModuleStatus + "</b></td></tr> "

            # html_body_string = html_body_string + testbody1 + testbody2 + testbody3 + testbody4 + testbody5

            self.numTestCaseFail = 0
            self.numTestCasePass = 0
        except Exception as e:
            print(f"'Report_Module' has this error -> \n {str(e)}")
        return testbody1 + testbody21 + testbody2 + testbody3 + testbody4 + testbody41 + testbody5

    def Report_ModuleSummary(self, browser, modulename, htmlBody, module_starttime_str):
        try:
            rstatus = "Pass"
            self.folders()

            # self.testmodulefilepath = browser + ".html"
            self.testmodulefilepath = "TestModules" + ".html"
            #self.Report_TestModule_Filename = str(self.TestModuleFolder) + "\\" + str(self.testmodulefilepath)
            self.Report_TestModule_Filename = os.path.join(str(self.TestModuleFolder), str(self.testmodulefilepath))

            if self.browserfound != browser or self.browserfound == "":
                self.browserfound = browser

                if not os.path.exists(self.Report_TestModule_Filename):
                    # make a new file if not
                    # if self.browserfound != browser or self.browserfound == "":
                    #   self.browserfound = browser

                    tw = open(self.Report_TestModule_Filename, 'w')

                    # tw.write(self.new_html_styles_part1)

                    tw.write("<html><head> ")
                    tw.write("<title>Selenium Python Automation</title>")
                    tw.write("<link rel=\"stylesheet\" href=\"../css/style.css\">")
                    tw.write("<script type=\"text/javascript\">")

                    # function to update the browser drop down values, called on page load inside body
                    tw.write("function updateTestStepOptions() {")
                    # tw.write("  // Get all the teststeps-divs divs")
                    tw.write("  const testStepDivs = document.querySelectorAll('.testsuites-divs');")
                    tw.write("console.log(\"testsuite count: \" + testStepDivs);")
                    # tw.write("  // Get the dropdown elements")
                    tw.write(
                        "  const browserDropdown = document.querySelector('select[name=\"test-browser-dropdown\"]');")

                    # tw.write("// Reset the dropdowns")
                    tw.write("  browserDropdown.innerHTML = '<option value=\"\">All</option>';")

                    # tw.write("  // Loop through each teststeps-divs div and extract the data-* attributes")
                    tw.write("  for (let i = 0; i < testStepDivs.length; i++) {")
                    tw.write("const div = testStepDivs[i];")
                    tw.write("const browser = div.getAttribute('data-browser');")

                    # tw.write("// Add the options to the dropdowns if they don't already exist")
                    tw.write("if (!browserDropdown.querySelector(`option[value=\"${browser}\"]`)) {")
                    tw.write("  const browserOption = document.createElement('option');")
                    tw.write("  browserOption.value = browser;")
                    tw.write("  browserOption.text = browser;")
                    tw.write("  browserDropdown.appendChild(browserOption);")
                    tw.write("}")
                    tw.write("}")
                    tw.write("}")

                    # function to filter data inside test suites page based on the drop down value
                    tw.write("function filterDivs() {")
                    tw.write("let noResult = 0;")
                    tw.write("const testStepDivs = document.querySelectorAll('.testsuites-divs');")
                    tw.write(
                        "const browserDropdown = document.querySelector('select[name=\"test-browser-dropdown\"]');")
                    tw.write("")
                    # tw.write("// Get the selected option values")
                    tw.write("const selectedBrowser = browserDropdown.options[browserDropdown.selectedIndex].value;")
                    tw.write("")
                    tw.write("for (let i = 0; i < testStepDivs.length; i++) {")
                    tw.write("const div = testStepDivs[i];")
                    tw.write("const browser = div.getAttribute('data-browser');")
                    tw.write("")
                    tw.write("let shouldDisplay = true;")
                    tw.write("if (selectedBrowser && selectedBrowser !== browser && selectedBrowser !== 'All') {")
                    tw.write("shouldDisplay = false;")
                    tw.write("}")
                    tw.write("")
                    tw.write("if (shouldDisplay) {")
                    tw.write("div.style.display = 'block';")
                    tw.write("noResult++;")
                    tw.write("} else {")
                    tw.write("div.style.display = 'none';")
                    tw.write("}")
                    tw.write("}")
                    tw.write("")
                    tw.write("console.log(\"noResult value is : \"+noResult);")
                    # tw.write("// Show a message if there are no results")
                    tw.write("const searchResultsMessage = document.querySelector('.search-results-message');")
                    tw.write("if (noResult === 0) {")
                    tw.write("searchResultsMessage.style.display = 'block';")
                    tw.write("} else {")
                    tw.write("searchResultsMessage.style.display = 'none';")
                    tw.write("}")
                    tw.write("}")
                    # ________________________________________

                    tw.write("</script>")
                    tw.write("</head>")
                    tw.write("<body onload=\"updateTestStepOptions()\">")
                    tw.write("<section id=\"sidebar\">")
                    tw.write("<div class=\"white-label\">")
                    tw.write("</div>")
                    tw.write("<div class=\"container\" id=\"sidebar-nav\">")

                    tw.write("<ul>")
                    tw.write(
                        "<li class=\"leftnav\"><a href=\"..\\TestReport_highReport.html\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                    tw.write(
                        "<li class=\"leftnav\"><a href=\"..\\Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                    tw.write(
                        "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                    tw.write(
                        "<li class=\"leftnav\"><a href=\"..\\TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")

                    tw.write(
                        "<li class=\"leftnav\"><a href=\"..\\AboutTheProject\\AboutTheProject.html\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                    tw.write(
                        "<li class=\"leftnav\"><a href=\"..\\About\\About.html\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                    tw.write("</ul>")

                    tw.write(self.new_html_styles_part3)

                    # Test Suites details below main header
                    tw.write("<div class=\"content-header\">")
                    tw.write("<h1> Test Suites </h1>")
                    tw.write("<p>Test Automation Report - Summary of Test Cases</p>")
                    tw.write("</div>")
                    tw.write("<br>")

                    # Adding drop down and button
                    tw.write(
                        "<select title=\"Select a browser to display\" class=\"browserfilter\" name=\"test-browser-dropdown\">")
                    tw.write("</select>")
                    tw.write("<button onclick=\"filterDivs()\">Search</button>")
                    tw.write("<div class=\"search-results-message\" style=\"display:none;\">")
                    tw.write("<p> There is no data with the search result</p>")
                    tw.write("</div>")

                    tw.write("<div data-browser=" + browser + " style=\"display:none;\" class=\"testsuites-divs\">")
                    tw.write("<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5>  Browser Name: "
                             + browser + "</h4>")
                    tw.write(
                        "<table class =\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100% align='center'><tr>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Module</b></td>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Description</b></td>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Total Test "
                        "Case(s) "
                        "</b></td>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Test Case("
                        "s) Passed "
                        "</b></td>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Test Case("
                        "s) Failed "
                        "</b></td>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Execution Time</b></td>")
                    tw.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Status</b></td>")

                    tw.write(" </tr>")

                    # print("htmlBody -> " + htmlBody)
                    tw.write(htmlBody)

                    tw.close()
                else:
                    tw1 = open(self.Report_TestModule_Filename, 'a')
                    tw1.write(
                        "</table> </div> <div data-browser=" + browser + " style=\"display:none;\" class=\"testsuites-divs\">")
                    tw1.write(
                        "<h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5>  Browser Name : "
                        + browser + "</h4>")
                    tw1.write(
                        "<table class =\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100% align='center'><tr>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Module</b></td>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Description</b></td>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Total Test "
                        "Case(s) "
                        "</b></td>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Test Case("
                        "s) Passed "
                        "</b></td>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Test Case("
                        "s) Failed "
                        "</b></td>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Execution Time</b></td>")
                    tw1.write(
                        "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                        "SIZE=2><b>Status</b></td>")

                    tw1.write(" </tr>")

                    # print("htmlBody -> " + htmlBody)
                    tw1.write(htmlBody)

                    tw1.close()
            else:
                rbw = open(self.Report_TestModule_Filename, 'a')

                # print("htmlBody rbw already exist -> " + htmlBody)
                rbw.write(htmlBody)

                rbw.close()
            self.detailedReportFail = 0
        except Exception as e:
            print(f"'Report_ModuleSummary' has this error ->\n {self.common_methods.error_message(e)}")

    def Report_Browser(self, browsername):
        try:
            # global html_body_string

            if self.numModuleFail > 0:
                BrowserStatus = "Fail"
                self.StatusDetbgcolor = self.Status_Color("fail")
            else:
                BrowserStatus = "Pass"
                self.StatusDetbgcolor = self.Status_Color("pass")

            # print("browsername and modulename -> " + browsername)

            if self.browser != browsername or self.browser == "":  # check this
                self.numTestCaseFail = 0
                self.numTestCasePass = 0
                self.browser = browsername

            # self.TotalModules = self.numModulePass + self.numModuleFail

            """
            # print("Report_TestCase_Filename in Report_Module -> " + Report_TestCase_Filename)
            testbody1 = "<tr><td width=20% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b><a href=" + \
                        self.Report_TestModule_Filename + ">" + browsername + "</a></b></td>"
    
            print("str(self.numModulePass+self.numModuleFail) -> " + str(self.numModulePass + self.numModuleFail))
            print("self.numModulePass -> " + str(self.numModulePass))
            print("self.numModuleFail) -> " + str(self.numModuleFail))
    
            testbody2 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                        + str(self.numModulePass + self.numModuleFail) + "</b></td>"
    
            testbody3 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                        + str(self.numModulePass) + "</b></td>"
    
            testbody4 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                        + str(self.numModuleFail) + "</b></td>"
    
            testbody5 = "<td width=15% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT COLOR=\"#FFFFFF\" " \
                                                                                            "FACE=\"Arial\" SIZE=2><b>" + \
                        BrowserStatus + "</b></td></tr> "
    
            # html_body_string = html_body_string + testbody1 + testbody2 + testbody3 + testbody4 + testbody5
    
            # self.numModuleFail = 0
            # self.numModulePass = 0
            return testbody1 + testbody2 + testbody3 + testbody4 + testbody5
            """
        except Exception as e:
            print(f"'Report_Browser' has this error -> \n {str(e)}")

    def Report_BrowserSummary(self, browsername, browser_startTime_str):
        try:
            # print(f"Report_TestModule_Filename is {self.Report_TestModule_Filename} and browser name is {browsername}")
            rstatus = "Pass"
            self.folders()

            if self.numModuleFail > 0:
                BrowserStatus = "Fail"
                self.StatusDetbgcolor = self.Status_Color("fail")
            else:
                BrowserStatus = "Pass"
                self.StatusDetbgcolor = self.Status_Color("pass")

            """
            a1 = "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
            "SIZE=2><b>Browser</b></td>"
    
            a2 = "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Total "
            "Module(s) </b></td>"
    
            a3 = "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Module("
            "s) Passed "
            "</b></td>"
    
            a4 = "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Module("
            "s) Failed "
            "</b></td>"
    
            a5 = "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
            "SIZE=2><b>Status</b></td>"
    
            a6 = " </tr>"
    
            a7 = "<tr><td width=20% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b><a href=" \
                 + self.Report_TestModule_Filename + ">" + browsername + "</a></b></td>"
    
            a8 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                 + str(self.TotalModules) + "</b></td>"
    
            a9 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                 + str(self.numModulePass) + "</b></td>"
    
            a10 = "<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                  + str(self.numModuleFail) + "</b></td>"
    
            a11 = "<td width=15% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT COLOR=\"#FFFFFF\" " \
                                                                                      "FACE=\"Arial\" SIZE=2><b>" + \
                  BrowserStatus + "</b></td></tr> "
    
            self.html_body_string = str(a1) + str(a2) + str(a3) + str(a4) + str(a5) + str(a6) + str(a7) + str(a8) + str(a9) \
                                    + str(a10) + str(a11)
            """
            browser_endTime = datetime.datetime.now()
            browser_endTime_str = browser_endTime.strftime("%d/%m/%Y %H:%M:%S")

            browser_start_datetime = datetime.datetime.strptime(browser_startTime_str, "%d/%m/%Y %H:%M:%S")
            browser_end_datetime = datetime.datetime.strptime(browser_endTime_str, "%d/%m/%Y %H:%M:%S")

            self.html_body_string = self.html_body_string + "<tr><td width=20% align=\"center\"><FONT COLOR=\"#153E7E\" " \
                                                            "FACE=\"Arial\" SIZE=2><b><a href=" + \
                                    self.Report_TestModule_Filename + ">" + browsername + "</a></b></td>" + \
                                    "<td width=20% align=\"center\"><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2><b> " + \
                                    str(browser_end_datetime - browser_start_datetime) + "</b></td>" + \
                                    "<td width=20% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT " \
                                                                                                        "COLOR=\"#153E7E" \
                                                                                                        "\" FACE=\"Arial" \
                                                                                                        "\" SIZE=2><b>" + \
                                    BrowserStatus + "</b></td></tr> "
            # print(f"self.html_body_string is - {self.html_body_string}")
        except Exception as e:
            print(f"'Report_BrowserSummary' has this error -> \n {str(e)}")

    def HighReport(self, projectName, UserReq, TestEnv, release, systemRun, overall_startTime):
        try:
            self.folders()
            common_methods = commonMethods()
            os_details = common_methods.get_os_details()

            # create css file
            self.writeCSS()

            successRate = 0.0
            failRate = 0.0

            passwidth = 0.0
            failwidth = 0.0

            # this is for the entire test cases combined
            NumTotal = self.FinalTestCasePass + self.FinalTestCaseFail
            if NumTotal != 0:
                successRate = self.FinalTestCasePass * 100 / NumTotal
                successRate = round(successRate, 2)

                failRate = 100 - successRate
                passwidth = (150 * successRate) / 100
                failwidth = 150 - passwidth

            # this if for individual modules/product vice test cases

            for product, values in self.test_cases_map.items():
                pass_value, fail_value, totalTC = values
                successRate = round((pass_value / totalTC) * 100)
                failRate = 100 - successRate
                passwidth = (150 * successRate) / 100
                failwidth = 150 - passwidth
                self.test_cases_map[product] += (successRate, passwidth, failRate, failwidth,)

            """
            if NumTotal != 0:
                successRate = self.FinalTestCasePass * 100 / NumTotal
                successRate = round(successRate, 2)

                failRate = 100 - successRate
                passwidth = (150 * successRate) / 100
                failwidth = 150 - passwidth
            """

            successRateMod = 0.0
            failRateMod = 0.0

            passwidthMod = 0.0
            failwidthMod = 0.0

            """
            NumTotalMod = self.numModulePass + self.numModuleFail
            if NumTotalMod != 0:
                successRateMod = self.numModulePass * 100 / NumTotalMod
                successRateMod = round(successRateMod, 2)
    
                failRateMod = 100 - successRateMod
                passwidthMod = (150 * successRateMod) / 100
                failwidthMod = 150 - passwidthMod
            """

            NumTotalMod = self.finalModulePass + self.finalModuleFail
            if NumTotalMod != 0:
                successRateMod = self.finalModulePass * 100 / NumTotalMod
                successRateMod = round(successRateMod, 2)

                failRateMod = 100 - successRateMod
                passwidthMod = (150 * successRateMod) / 100
                failwidthMod = 150 - passwidthMod

            # this if for individual modules/product vice sub modules

            for product, values in self.test_modules_map.items():
                pass_value1, fail_value1, totalTC1 = values
                successRate1 = round((pass_value1 / totalTC1) * 100)
                failRate1 = 100 - successRate1
                passwidth1 = (150 * successRate1) / 100
                failwidth1 = 150 - passwidth1
                self.test_modules_map[product] += (successRate1, passwidth1, failRate1, failwidth1,)
                print(f"product inside test_modulesmap is {product}")

            print(f"test case map data : {self.test_cases_map}")
            print(f"test module map data : {self.test_modules_map}")

            header = "<html><head> " + \
                     "<title>Selenium Python Automation</title>" + \
                     "<link rel=\"stylesheet\" href=\"css/style.css\">" + \
                     "</head>" + \
                     "<body>" + \
                     "<section id=\"sidebar\">" + \
                     "<div class=\"white-label\">" + \
                     "</div>" + \
                     "<div class=\"container\" id=\"sidebar-nav\">"
            new_html_styles_part2_li = ""
            new_html_styles_part3 = "</div>" + \
                                    "</section>" + \
                                    "<section id=\"content-div\">" + \
                                    "<div id=\"header\">" + \
                                    "<div class=\"header-nav\">" + \
                                    "<div class=\"menu-button\">Test Automation Report - " + projectName + "</div>" + \
                                    "<div class=\"nav\">" + \
                                    "<ul>" + \
                                    "<li class=\"nav-profile\">" + \
                                    "<div class=\"nav-profile-name\">" + \
                                    os.getlogin() + "<i class=\"fa fa-caret-down\"></i>" + \
                                    "</div>" + \
                                    "</li>" + \
                                    "</ul>" + \
                                    "</div>" + \
                                    "</div>" + \
                                    "</div>" + \
                                    "<div class=\"content\" id=\"dashboard\" style=\"display: block;\">"

            html_header_string = "<br><div class=\"topLeft\"><span></span> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> " \
                                 "<h4>Test Details :</h4> </FONT><br>" + \
                                 "<table class=\"TestDetails\" width=\"100%\" border=1 cellspacing=1 " \
                                 "cellpadding=1><tr><td bgcolor=\"#153E7E\" style=\"padding-top:-30px;\">" + \
                                 "<FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Run Date</b></td><td><FONT " \
                                 "COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + \
                                 self.dt_string + "</b></td></tr>"

            if not os.path.exists(self.highReportWithGraph):
                # make a new file if not
                tw = open(self.highReportWithGraph, 'w')

                # tw.write(header)

                tw.write("<html><head> ")
                tw.write("<title>Selenium Python Automation</title>")
                tw.write("<link rel=\"stylesheet\" href=\"css/style.css\">")
                tw.write("<script type=\"text/javascript\">")

                # Product Details tab function_____________________________
                tw.write("function initTabs(){")
                tw.write("const tabs = document.querySelectorAll('.tabs a');")
                tw.write("const tabContent = document.querySelectorAll('.tab-content');")
                tw.write("")
                tw.write("tabs.forEach(tab => {")
                tw.write("tab.addEventListener('click', e => {")
                tw.write("e.preventDefault();")
                tw.write("const tabId = tab.getAttribute('href');")
                tw.write("")
                tw.write("tabs.forEach(tab => {")
                tw.write("tab.classList.remove('active');")
                tw.write("});")
                tw.write("tabContent.forEach(content => {")
                tw.write("content.classList.remove('active');")
                tw.write("});")
                tw.write("")
                tw.write("tab.classList.add('active');")
                # tw.write("document.querySelector(tabId).classList.add('active');")
                tw.write("const divsToActive = document.querySelectorAll(tabId);")
                tw.write("divsToActive.forEach(div => {")
                tw.write("div.classList.add('active');")
                tw.write("});")
                tw.write("});")
                tw.write("});")
                tw.write("tabContent[0].classList.add('active');")
                tw.write("}")
                tw.write("initTabs();")

                # code to update widget header for Test Cases and Test Modules
                tw.write("function openTab(evt, tabName){")
                tw.write("const headerTitle1 = document.getElementById(\"header-title-testcases\");")
                tw.write("headerTitle1.textContent = \"Test Cases Summary - \" + evt.target.textContent;")
                tw.write("const headerTitle2 = document.getElementById(\"header-title-testmodule\");")
                tw.write("headerTitle2.textContent = \"Test Modules Summary - \" + evt.target.textContent;")
                tw.write("}")

                # __________________________________

                tw.write("</script>")
                tw.write("</head>")
                tw.write("<body onload=\"initTabs()\">")
                tw.write("<section id=\"sidebar\">")
                tw.write("<div class=\"white-label\">")
                tw.write("</div>")
                tw.write("<div class=\"container\" id=\"sidebar-nav\">")

                tw.write("<ul>")
                tw.write(
                    "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"AboutTheProject\\AboutTheProject.html\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"About\\About.html\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                tw.write("</ul>")

                tw.write(new_html_styles_part3)

                tw.write("<div class=\"content-header\">")
                tw.write("<h1>Dashboard</h1>")
                tw.write("<p>Test Automation Report - Summary of the Test Run</p>")
                tw.write("</div>")

                # Test Details Widget ______________________________________________
                tw.write("<div class=\"widget-box sample-widget\">")
                tw.write("<div class=\"widget-header\">")
                tw.write(" <h2>Test Details</h2>")
                tw.write("<i class=\"fa fa-cog\"></i>")
                tw.write("</div>")
                tw.write("<div class=\"widget-content\" style=\"padding:13px;\">")
                tw.write("<div class=\"summary-container\">")

                tw.write("<table class=\"TestDetails\" width=\"100%\" cellspacing=\"1\" style=\"height:261px;\">")
                tw.write(" <tbody>")
                tw.write("<tr>")
                tw.write("<td bgcolor=\"#0274BD\" style=\"padding-top:-30px;\">")
                tw.write("<font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\">")
                tw.write("<b>Run Date</b></font>")
                tw.write("</td>")
                tw.write("<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\">")
                tw.write(
                    "<b>" + self.formatted_DDMMMYYYY + " - " + self.formatted_Time_now + " </b>")

                tw.write("<b><div class=\"tooltip\">")
                tw.write(''.join([i[0] for i in self.timezone.split()]))
                tw.write("<div class=\"tooltiptext\"> " + self.timezone + "</div>")
                tw.write("</div></b>")
                tw.write("</font>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>User Requested</b></font></td>")
                tw.write("<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + str(
                    self.common_methods.getusername()) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>Environment</b></font></td>")
                tw.write("<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + TestEnv + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>Test Start Time</b></font></td>")

                overall_startTime_str = overall_startTime.strftime("%d/%m/%Y %H:%M:%S")
                datetime_obj = datetime.datetime.strptime(overall_startTime_str, "%d/%m/%Y %H:%M:%S")
                new_starttime = datetime_obj.strftime("%d %B %Y %I:%M:%S %p")

                tw.write("<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + str(
                    new_starttime) + " </b>")
                tw.write("<b><div class=\"tooltip\">")
                tw.write(''.join([i[0] for i in self.timezone.split()]))
                tw.write("<div class=\"tooltiptext\"> " + self.timezone + "</div>")
                tw.write("</div></b>")
                tw.write("</font></td>")
                tw.write("</tr>")
                tw.write("<tr>")

                endTime = datetime.datetime.now()
                endTime_str = endTime.strftime("%d/%m/%Y %H:%M:%S")
                datetime_obj = datetime.datetime.strptime(endTime_str, "%d/%m/%Y %H:%M:%S")
                new_endtime = datetime_obj.strftime("%d %B %Y %I:%M:%S %p")

                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>Test End Time</b></font></td>")
                tw.write(
                    "<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + new_endtime + " </b>")
                tw.write("<b><div class=\"tooltip\">")
                tw.write(''.join([i[0] for i in self.timezone.split()]))
                tw.write("<div class=\"tooltiptext\"> " + self.timezone + "</div>")
                tw.write("</div></b>")
                tw.write("</font></td>")
                tw.write("</tr>")
                tw.write("<tr>")

                start_datetime = datetime.datetime.strptime(overall_startTime_str, "%d/%m/%Y %H:%M:%S")
                end_datetime = datetime.datetime.strptime(endTime_str, "%d/%m/%Y %H:%M:%S")
                ranTime = end_datetime - start_datetime
                # datetime_obj = datetime.datetime.strptime(str(ranTime), "%d/%m/%Y %H:%M:%S")
                # new_ranTime = datetime_obj.strftime("%d %B %Y %I:%M:%S %p")

                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>Total Test Ran Time</b></font></td>")
                tw.write(
                    "<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + str(ranTime) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>Release</b></font></td>")
                tw.write(
                    "<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + str(release) + "</b></font></td>")
                tw.write("</tr>")

                tw.write("<tr>")
                tw.write(
                    "<td bgcolor=\"#0274BD\"><font color=\"#E0E0E0\" face=\"Arial\" size=\"2.75\"><b>Operating System</b></font></td>")
                tw.write(
                    "<td><font color=\"#153E7E\" face=\"Arial\" size=\"2.75\"><b>" + str(os_details) + "</b></font></td>")
                tw.write("</tr>")

                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</div>")

                tw.write("</div>")
                tw.write("</div>")

                # Test Module Summary Widget ______________________________________________
                tw.write("<div class=\"widget-box sample-widget\">")
                tw.write("<div class=\"widget-header\">")
                tw.write("<h2 id=\"header-title-testmodule\">Test Module Summary</h2>")
                tw.write("<i class=\"fa fa-cog\"></i>")
                tw.write("</div>")
                tw.write("<div class=\"widget-content\" style=\"padding:31px;\">")

                # default module data
                tw.write("<div id=\"" + self.productname_ProductDetails_sheet + "\" class=\"tab-content active\">")
                tw.write(
                    "<table class=\"t06parent\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"height:261px;\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write("<td width=\"50%\">")
                tw.write("<table class=\"t06\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"50\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Modules</b></font></td>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                        NumTotalMod) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Pass</b></font></td>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                        self.finalModulePass) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Fail</b></font></td>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                        self.finalModuleFail) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</td>")
                tw.write("<td width=\"50%\">")
                tw.write("<font color=\"660000\" face=\"Arial\" size=\"3.5\">")
                tw.write(
                    "<table class=\"ModulePerTable\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"border-left: 1px solid grey;\">")
                tw.write("<tbody>")
                tw.write("<tr style=\"vertical-align: bottom;\">")
                tw.write("<td width=\"50%\">")
                tw.write("<table class=\"t07\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write("<td align=\"center\" valign=\"bottom\" style=\"padding:10px;\">")
                tw.write("<b>" + str(round(successRateMod, 2)) + " %</b>")
                # tw.write(
                #    "<div class=\"Passdiv\" style=\"height: 0.0px; background-color:green\"><span class=\"PercentageBarFont\"> 0</span></div>")
                tw.write("<div class=\"Passdiv\" style=\"width:50px; height: " + str(passwidthMod) +
                         "px; background-color:green\"><span class=\"PercentageBarFont\"> " + str(
                    self.finalModulePass) + "</span></div>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td style=\"height:1px; padding:10px;\" align=\"center\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Pass</b></font></td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</td>")
                tw.write("<td width=\"50%\">")
                tw.write("<table class=\"t08\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write("<td align=\"center\" valign=\"bottom\" style=\"padding:10px;\">")
                tw.write("<b>" + str(round(failRateMod, 2)) + " %</b>")
                # tw.write(
                #    "<div class=\"Faildiv\" style=\"height: 150.0px; background-color:red \"><span class=\"PercentageBarFont\">1</span></div>")
                tw.write("<div class=\"Faildiv\" style=\"width:50px;height: " + str(
                    failwidthMod) + "px; background-color:red\"><span class=\"PercentageBarFont\">" + str(
                    self.finalModuleFail) +
                         "</span></div>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td style=\"height:1px; padding:10px;\" align=\"center\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Fail</b></font></td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</font>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</div>")

                #  _______________________________
                # _Generate for each module test case section ____________________
                for product11, values11 in self.test_modules_map.items():
                    pass_value11, fail_value11, totalMods11, successRate11, passwidth11, failRate11, failwidth11 = values11
                    tw.write("<div id=" + str(product11) + " class=\"tab-content\">")

                    tw.write(
                        "<table class=\"t06parent\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"height:261px;\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write("<td width=\"50%\">")
                    tw.write("<table class=\"t06\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"50\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Modules</b></font></td>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                            totalMods11) + "</b></font></td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Pass</b></font></td>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                            pass_value11) + "</b></font></td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Fail</b></font></td>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                            fail_value11) + "</b></font></td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</td>")
                    tw.write("<td width=\"50%\">")
                    tw.write("<font color=\"660000\" face=\"Arial\" size=\"3.5\">")
                    tw.write(
                        "<table class=\"ModulePerTable\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"border-left: 1px solid grey;\">")
                    tw.write("<tbody>")
                    tw.write("<tr style=\"vertical-align: bottom;\">")
                    tw.write("<td width=\"50%\">")
                    tw.write("<table class=\"t07\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write("<td align=\"center\" valign=\"bottom\" style=\"padding:10px;\">")
                    tw.write("<b>" + str(round(successRate11, 2)) + " %</b>")
                    # tw.write(
                    #    "<div class=\"Passdiv\" style=\"height: 0.0px; background-color:green\"><span class=\"PercentageBarFont\"> 0</span></div>")
                    tw.write("<div class=\"Passdiv\" style=\"width:50px;height: " + str(passwidth11) +
                             "px; background-color:green\"><span class=\"PercentageBarFont\"> " + str(
                        pass_value11) + "</span></div>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td style=\"height:1px; padding:10px;\" align=\"center\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Pass</b></font></td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</td>")
                    tw.write("<td width=\"50%\">")
                    tw.write("<table class=\"t08\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write("<td align=\"center\" valign=\"bottom\" style=\"padding:10px;\">")
                    tw.write("<b>" + str(round(failRate11, 2)) + " %</b>")
                    # tw.write(
                    #    "<div class=\"Faildiv\" style=\"height: 150.0px; background-color:red \"><span class=\"PercentageBarFont\">1</span></div>")
                    tw.write("<div class=\"Faildiv\" style=\"width:50px;height: " + str(
                        failwidth11) + "px; background-color:red\"><span class=\"PercentageBarFont\">" + str(
                        fail_value11) +
                             "</span></div>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td style=\"height:1px; padding:10px;\" align=\"center\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Fail</b></font></td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</font>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</div>")
                # ______________________________________________________________

                tw.write("</div>")
                tw.write("</div>")

                # product Details widget ____________________________________
                tw.write("<div class=\"widget-box sample-widget\">")
                tw.write("<div class=\"widget-header\">")
                tw.write("<h2>Product Details</h2>")
                tw.write("<i class=\"fa fa-cog\"></i>")
                tw.write("</div>")
                tw.write("<div class=\"widget-content\" style=\"padding:15px;\">")
                tw.write("<div class=\"tabbed-widget-container\">")
                tw.write("<div class=\"tabbed-widget\" style=\"height:239px;\">")
                tw.write("<ul class=\"tabs\">")
                # tw.write("<li><a href=\"#All\">All Modules</a></li>")

                """
                for productExcel in productDetails.iter_rows(values_only=True):
                    product = productExcel

                    productname = product.__getitem__(0)
                    product_details_columndata = product.__getitem__(1)
                    product_display = product.__getitem__(2)
                    if product_display == 'Y':
                        tw.write(
                            "<li><a href=\"#" + productname + "\"  onclick=\"openTab(event, ' " + productname + "')\">" + productname + "</a></li>")

                tw.write("</ul>")

                for productExcel in productDetails.iter_rows(values_only=True):
                    product = productExcel

                    productname = product.__getitem__(0)
                    product_details_columndata = product.__getitem__(1)
                    product_display = product.__getitem__(2)
                    if product_display == 'Y':
                        tw.write("<div id=\"" + productname + "\" class=\"tab-content\">")
                        # tw.write("<summary>Details 1</summary>")
                        tw.write("<p>" + product_details_columndata + "</p>")
                        tw.write("</div>")
                """

                tw.write(
                    "<li><a href=\"#" + self.productname_ProductDetails_sheet + "\"  onclick=\"openTab(event, ' " + self.productname_ProductDetails_sheet + "')\">" + self.productname_ProductDetails_sheet + "</a></li>")


                """
                for row in modules_sheet.iter_rows(min_row=2, values_only=True):
                    run_column_values = row[3]
                    main_modules_column_value = row[0]
                    if run_column_values == 'Y' or run_column_values == 'y':
                        if self.report_modulename != main_modules_column_value or self.report_modulename == "":
                            self.report_modulename = main_modules_column_value
                            productDetails = workbook["ProductDetails"]
                            for pd_row in productDetails.iter_rows(min_row=2, values_only=True):
                                product_name_column_value = pd_row[0]
                                details_column_value = pd_row[1]
                                if product_name_column_value == main_modules_column_value:
                                    tw.write(
                                        "<li><a href=\"#" + main_modules_column_value + "\"  onclick=\"openTab(event, ' " + main_modules_column_value + "')\">" + main_modules_column_value + "</a></li>")
                """

                for product, values in self.test_modules_map.items():
                    # productDetails = workbook["ProductDetails"]
                    for pd_row in self.productDetails.iter_rows(min_row=2, values_only=True):
                        product_name_column_value = pd_row[0]
                        details_column_value = pd_row[1]
                        if product_name_column_value == product:
                            tw.write(
                                "<li><a href=\"#" + product + "\"  onclick=\"openTab(event, ' " + product + "')\">" + product + "</a></li>")

                tw.write("</ul>")

                tw.write("<div id=\"" + self.productname_ProductDetails_sheet + "\" class=\"tab-content\">")
                # tw.write("<summary>Details 1</summary>")
                tw.write("<p>" + self.product_details_columndata + "</p>")
                tw.write("</div>")

                """
                self.report_modulename = ""
                for row in modules_sheet.iter_rows(min_row=2, values_only=True):
                    run_column_values = row[3]
                    main_modules_column_value = row[0]
                    if run_column_values.lower() == 'y':
                        if self.report_modulename != main_modules_column_value or self.report_modulename == "":
                            self.report_modulename = main_modules_column_value
                            productDetails = workbook["ProductDetails"]
                            for pd_row in productDetails.iter_rows(min_row=2, values_only=True):
                                product_name_column_value = pd_row[0]
                                details_column_value = pd_row[1]
                                if product_name_column_value.lower() == main_modules_column_value.lower():
                                    tw.write("<div id=\"" + main_modules_column_value + "\" class=\"tab-content\">")
                                    # tw.write("<summary>Details 1</summary>")
                                    tw.write("<p>" + details_column_value + "</p>")
                                    tw.write("</div>")
                """

                for product, values in self.test_modules_map.items():
                    # productDetails = workbook["ProductDetails"]
                    for pd_row in self.productDetails.iter_rows(min_row=2, values_only=True):
                        product_name_column_value = pd_row[0]
                        details_column_value = pd_row[1]
                        if product_name_column_value.lower() == product.lower():
                            tw.write("<div id=\"" + product + "\" class=\"tab-content\">")
                            # tw.write("<summary>Details 1</summary>")
                            tw.write("<p>" + details_column_value + "</p>")
                            tw.write("</div>")

                tw.write("</div>")
                tw.write("</div>")
                tw.write("</div>")
                tw.write("</div>")
                # _________________________________________

                # Test Cases Summary Widget _________________________________________

                tw.write("<div class=\"widget-box sample-widget\">")
                tw.write("<div class=\"widget-header\">")
                tw.write("<h2 id=\"header-title-testcases\">Test Cases Summary</h2>")
                tw.write("<i class=\"fa fa-cog\"></i>")
                tw.write("</div>")
                tw.write("<div class=\"widget-content\" style=\"padding:15px; height:260px;\">")

                # default test cases number for "All" ______________________________
                tw.write("<div id=\"" + self.productname_ProductDetails_sheet + "\" class=\"tab-content active\">")
                tw.write(
                    "<table class=\"t06parent\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"height:233px;\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write("<td width=\"50%\">")
                tw.write("<table class=\"t06\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"50\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Test Cases</b></font></td>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                        NumTotal) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Pass</b></font></td>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                        self.FinalTestCasePass) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Fail</b></font></td>")
                tw.write(
                    "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">:  " + str(
                        self.FinalTestCaseFail) + "</b></font></td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</td>")
                tw.write("<td width=\"50%\">")
                tw.write("<font color=\"660000\" face=\"Arial\" size=\"3.5\">")
                tw.write(
                    "<table class=\"ModulePerTable\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"border-left: 1px solid grey;\">")
                tw.write("<tbody>")
                tw.write("<tr style=\"vertical-align: bottom;\">")
                tw.write("<td width=\"50%\">")
                tw.write("<table class=\"t07\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write("<td align=\"center\" valign=\"bottom\" style=\"padding:10px;\">")
                tw.write("<b>" + str(round(successRate, 2)) + "%</b>")
                # tw.write(
                #    "<div class=\"Passdiv\" style=\"height: 100.0px; background-color:green\"><span class=\"PercentageBarFont\" color=\"white\"> 2</span></div>")
                tw.write("<div class=\"Passdiv\" style=\"width:50px;height: " + str(
                    passwidth) + "px; background-color:green\"><span class=\"PercentageBarFont\">" + str(
                    self.FinalTestCasePass) + "</span></div>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td style=\"height: 3px; padding:10px;\" align=\"center\" cellspacing=\"4\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Pass</b></font></td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</td>")
                tw.write("<td width=\"50%\">")
                tw.write("<table class=\"t08\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                tw.write("<tbody>")
                tw.write("<tr>")
                tw.write("<td align=\"center\" valign=\"bottom\">")  # style=\"padding:10px;\">")
                tw.write("<b>" + str(round(failRate, 2)) + " %</b>")
                # tw.write(
                #    "<div class=\"Faildiv\" style=\"height: 49.95px; background-color:red \"><span class=\"PercentageBarFont\">1</span></div>")
                tw.write("<div class=\"Faildiv\" style=\"width:50px;height: " + str(
                    failwidth) + "px; background-color:red\"><span class=\"PercentageBarFont\">" + str(
                    self.FinalTestCaseFail) + "</span></div>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("<tr>")
                tw.write(
                    "<td style=\"height:1px; padding:10px\" align=\"center\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Fail</b></font></td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</font>")
                tw.write("</td>")
                tw.write("</tr>")
                tw.write("</tbody>")
                tw.write("</table>")
                tw.write("</div>")
                # _________________________________________________

                # _Generate for each module test case section ____________________
                for product, values in self.test_cases_map.items():
                    pass_value, fail_value, totalTC, successRate, passwidth, failRate, failwidth = values
                    tw.write("<div id=" + str(product) + " class=\"tab-content\">")
                    tw.write(
                        "<table class=\"t06parent\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"height:233px;\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write("<td width=\"50%\">")
                    tw.write("<table class=\"t06\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"50\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Test Cases</b></font></td>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                            totalTC) + "</b></font></td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Pass</b></font></td>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">: " + str(
                            pass_value) + "</b></font></td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Total Fail</b></font></td>")
                    tw.write(
                        "<td align=\"left\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">:  " + str(
                            fail_value) + "</b></font></td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</td>")
                    tw.write("<td width=\"50%\">")
                    tw.write("<font color=\"660000\" face=\"Arial\" size=\"3.5\">")
                    tw.write(
                        "<table class=\"ModulePerTable\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\" width=\"100%\" style=\"border-left: 1px solid grey;\">")
                    tw.write("<tbody>")
                    tw.write("<tr style=\"vertical-align: bottom;\">")
                    tw.write("<td width=\"50%\">")
                    tw.write("<table class=\"t07\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write("<td align=\"center\" valign=\"bottom\" style=\"padding:10px;\">")
                    tw.write("<b>" + str(round(successRate, 2)) + "%</b>")
                    # tw.write(
                    #    "<div class=\"Passdiv\" style=\"height: 100.0px; background-color:green\"><span class=\"PercentageBarFont\" color=\"white\"> 2</span></div>")
                    tw.write("<div class=\"Passdiv\" style=\"width:50px;height: " + str(
                        passwidth) + "px; background-color:green\"><span class=\"PercentageBarFont\">" + str(
                        pass_value) + "</span></div>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td style=\"height: 3px; padding:10px;\" align=\"center\" cellspacing=\"4\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Pass</b></font></td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</td>")
                    tw.write("<td width=\"50%\">")
                    tw.write("<table class=\"t08\" border=\"0\" cellspacing=\"1\" cellpadding=\"1\">")
                    tw.write("<tbody>")
                    tw.write("<tr>")
                    tw.write("<td align=\"center\" valign=\"bottom\">")  # style=\"padding:10px;\">")
                    tw.write("<b>" + str(round(failRate, 2)) + " %</b>")
                    # tw.write(
                    #    "<div class=\"Faildiv\" style=\"height: 49.95px; background-color:red \"><span class=\"PercentageBarFont\">1</span></div>")
                    tw.write("<div class=\"Faildiv\" style=\"width:50px;height: " + str(
                        failwidth) + "px; background-color:red\"><span class=\"PercentageBarFont\">" + str(
                        fail_value) + "</span></div>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("<tr>")
                    tw.write(
                        "<td style=\"height:1px; padding:10px\" align=\"center\"><font color=\"#000066\" face=\"Arial\" size=\"2.75\"><b style=\"white-space:nowrap;\">Fail</b></font></td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</font>")
                    tw.write("</td>")
                    tw.write("</tr>")
                    tw.write("</tbody>")
                    tw.write("</table>")
                    tw.write("</div>")

                # ________________________________________________________________
                tw.write("</div>")
                tw.write("</div>")
                # ---------------------------------------------------

                tw.write("</div>")
                tw.write("</body></html>")

                """
                tw.write(html_header_string)
    
                tw.write(
                    "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>User "
                    "Requested</b></td>")
                tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + UserReq + "</b></td></tr>")
    
                tw.write(
                    " <tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Environment</b></td>")
                tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + TestEnv + "</b></td></tr>")
    
                tw.write(
                    "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Test Start "
                    "Time</b></td>")
    
                overall_startTime_str = overall_startTime.strftime("%d/%m/%Y %H:%M:%S")
    
                tw.write(
                    " <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + str(
                        overall_startTime_str) + "</b></td></tr>")
    
                endTime = datetime.datetime.now()
                endTime_str = endTime.strftime("%d/%m/%Y %H:%M:%S")
    
                tw.write(
                    "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Test End "
                    "Time</b></td>")
                tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + str(endTime_str) + "</b></td></tr>")
    
                start_datetime = datetime.datetime.strptime(overall_startTime_str, "%d/%m/%Y %H:%M:%S")
                end_datetime = datetime.datetime.strptime(endTime_str, "%d/%m/%Y %H:%M:%S")
    
                tw.write(
                    "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Total Test Ran "
                    "Time</b></td>")
                tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + str(end_datetime - start_datetime) +
                         "</b></td></tr>")
    
                tw.write(
                    " <tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Release</b></td>")
                tw.write(
                    " <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + release + "</b></td></tr></table></div>")
    
                tw.write(
                    "<div class=\"topRight\"><h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Test Result Summary "
                    ":</h4> <br><br>")
    
                tw.write("<table class =\"TestDetails\">")
                tw.write("<tr>")
                tw.write(
                    "<td width=50% height=\"234px\" style=\"padding: 3px; border-right: double 2px lightgrey; "
                    "vertical-align:text-top; \" align=\"center\">")
    
                tw.write(
                    "<h3 class=\"heading\"> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=3.5> Test Modules Summary</h3>")
    
                tw.write("<table class=\"t06\" border=0 cellspacing=1 cellpadding=1 width=\"50\">")
    
                tw.write(
                    "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Total Modules</b></td>")
                tw.write(
                    "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">: ")
                tw.write(str(self.TotalModules) + "</b></td></tr>")
                tw.write(
                    "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Total Pass</b></td>")
                tw.write(
                    "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">: ")
                tw.write(str(self.numModulePass) + "</b></td></tr>")
    
                tw.write(
                    "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Total Fail</b></td>")
                tw.write(
                    "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">:  ")
                tw.write(str(self.numModuleFail) + "</b></td></tr></table>")
    
                tw.write(
                    " <table class=\"t07 td02\" border=0 cellspacing=1 cellpadding=1><tr><td valign=\"bottom\"><b>" +
                    str(successRateMod) + "%</b><div class=\"Passdiv\" style=\"height: " + str(passwidthMod) +
                    "px; \"><span class=\"PercentageBarFont\"> " + str(self.numModulePass) + "</span></div></td></tr>")
    
                tw.write(
                    "<tr><td style=\"height:1px;\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Pass</b></td></tr></table>")
    
                tw.write(
                    "<table class=\"t08\" border=0 cellspacing=1 cellpadding=1><tr><td align = \"center\" "
                    "valign=\"bottom\"><b>" + str(failRateMod) + "%</b><div class=\"Faildiv\" style=\"height: " + str(
                        failwidthMod) + "px; \"><span class=\"PercentageBarFont\">" + str(self.numModuleFail) +
                    "</span></div></td></tr>")
                tw.write(
                    "<tr><td style=\"height:1px;\" align=\"center\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Fail</b></td></tr></table>")
    
                tw.write("</td>")
    
                tw.write(
                    "<td width=50% style=\"padding: 3px; vertical-align:text-top;\" align=\"center\"><h3 "
                    "class=\"heading\"> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=3.5> Test Cases Summary</h3>")
                tw.write(
                    "<table class=\"t06\" border=0 cellspacing=1 cellpadding=1 width=\"50\"><tr><td align=\"left\"><FONT "
                    "COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b style=\"white-space:nowrap;\">Total Cases</b></td>")
                tw.write(
                    "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">: " + str(NumTotal) + "</b></td></tr>")
                tw.write(
                    "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Total Pass</b></td>")
                tw.write(
                    "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">: " + str(self.FinalTestCasePass) + "</b></td></tr>")
                tw.write(
                    "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Total Fail</b></td>")
                tw.write(
                    "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">: " + str(self.FinalTestCaseFail) + "</b></td></tr></table>")
                tw.write(
                    "<table class=\"td02\" border=0 cellspacing=1 cellpadding=1><tr><td valign=\"bottom\" "
                    "align=\"center\"><b>" + str(successRate) + "%</b>")
                tw.write(
                    "<div class=\"Passdiv\" style=\"height: " + str(
                        passwidth) + "px; \"><span class=\"PercentageBarFont\">" + str(
                        self.FinalTestCasePass) + "</span></div></td></tr>")
                tw.write(
                    "<tr><td valign=\"bottom\" style=\"height: 1px; \" align=\"center\"><FONT COLOR=\"#000066\" "
                    "FACE=\"Arial\" SIZE=2.75><b style=\"white-space:nowrap;\">Pass</b></td></tr></table>")
                tw.write(
                    "<table class=\"t08\" border=0 cellspacing=1 cellpadding=1><tr><td valign=\"bottom\" "
                    "align=\"center\"><b>" + str(failRate) + "%</b>")
                tw.write(
                    "<div class=\"Faildiv\" style=\"height: " + str(
                        failwidth) + "px; \"><span class=\"PercentageBarFont\">" + str(
                        self.FinalTestCaseFail) + "</span></div></td></tr>")
                tw.write(
                    "<tr><td valign=\"bottom\" style=\"height: 1px; \" align=\"center\"><FONT COLOR=\"#000066\" "
                    "FACE=\"Arial\" SIZE=2.75><b "
                    "style=\"white-space:nowrap;\">Fail</b></td></tr></table></td></tr></table></div>")
    
                tw.write(
                    "<div class=\"TopBottom\"> <h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5>  Detailed Test Report "
                    ":</h4>")
    
                tw.write(
                    "<table class =\"DetailedReport\" border=1 cellspacing=1 cellpadding=1 width=100% align='center'>")  # <tr>")
    
                tw.write(
                    "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                    "SIZE=2><b>Browser</b></td>")
    
                tw.write(
                    "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Execution "
                    "Time</b></td>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                    "SIZE=2><b>Status</b></td>")
    
                tw.write(" </tr>")
    
                tw.write(self.html_body_string)
    
                tw.write("</table></div></body></html>")
                """
                tw.close()
                self.about_the_project(projectName)
                self.about_report(projectName)
        except Exception as e:
            print(e)
            print(f"'HighReport' has this error -> \n {self.common_methods.error_message(e)}")

    def HighReport_olddashboard(self, projectName, UserReq, TestEnv, release, systemRun, overall_startTime):
        self.folders()

        # create css file
        self.writeCSS()

        successRate = 0.0
        failRate = 0.0

        passwidth = 0.0
        failwidth = 0.0

        NumTotal = self.FinalTestCasePass + self.FinalTestCaseFail
        if NumTotal != 0:
            successRate = self.FinalTestCasePass * 100 / NumTotal
            successRate = round(successRate, 2)

            failRate = 100 - successRate
            passwidth = (150 * successRate) / 100
            failwidth = 150 - passwidth

        successRateMod = 0.0
        failRateMod = 0.0

        passwidthMod = 0.0
        failwidthMod = 0.0

        NumTotalMod = self.numModulePass + self.numModuleFail
        if NumTotalMod != 0:
            successRateMod = self.numModulePass * 100 / NumTotalMod
            successRateMod = round(successRateMod, 2)

            failRateMod = 100 - successRateMod
            passwidthMod = (150 * successRateMod) / 100
            failwidthMod = 150 - passwidthMod

        header = "<html><head> " + \
                 "<title>Selenium Python Automation</title>" + \
                 "<link rel=\"stylesheet\" href=\"css/style.css\">" + \
                 "</head>" + \
                 "<body>" + \
                 "<section id=\"sidebar\">" + \
                 "<div class=\"white-label\">" + \
                 "</div>" + \
                 "<div class=\"container\" id=\"sidebar-nav\">"
        new_html_styles_part2_li = ""
        new_html_styles_part3 = "</div>" + \
                                "</section>" + \
                                "<section id=\"content-div\">" + \
                                "<div id=\"header\">" + \
                                "<div class=\"header-nav\">" + \
                                "<div class=\"menu-button\">Test Automation Report - " + projectName + "</div>" + \
                                "<div class=\"nav\">" + \
                                "<ul>" + \
                                "<li class=\"nav-profile\">" + \
                                "<div class=\"nav-profile-name\">" + \
                                os.getlogin() + "<i class=\"fa fa-caret-down\"></i>" + \
                                "</div>" + \
                                "</li>" + \
                                "</ul>" + \
                                "</div>" + \
                                "</div>" + \
                                "</div>" + \
                                "<div class=\"content\" id=\"dashboard\" style=\"display: block;\">"

        html_header_string = "<br><div class=\"topLeft\"><span></span> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> " \
                             "<h4>Test Details :</h4> </FONT><br>" + \
                             "<table class=\"TestDetails\" width=\"100%\" border=1 cellspacing=1 " \
                             "cellpadding=1><tr><td bgcolor=\"#153E7E\" style=\"padding-top:-30px;\">" + \
                             "<FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Run Date</b></td><td><FONT " \
                             "COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + \
                             self.dt_string + "</b></td></tr>"

        if not os.path.exists(self.highReportWithGraph):
            # make a new file if not
            tw = open(self.highReportWithGraph, 'w')

            tw.write(header)

            tw.write("<ul>")
            tw.write(
                "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
            tw.write(
                "<li class=\"leftnav\"><a href=\"Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
            tw.write(
                "<li class=\"leftnav\"><a href=\"TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
            tw.write(
                "<li class=\"leftnav\"><a href=\"TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")
            tw.write(
                "<li class=\"leftnav\"><a href=\"#\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
            tw.write("<li class=\"leftnav\"><a href=\"#\" class=\"toggle-btn\" data-content=\"about\">About</a></li>")
            tw.write("</ul>")

            tw.write(new_html_styles_part3)
            tw.write(html_header_string)

            tw.write(
                "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>User "
                "Requested</b></td>")
            tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + UserReq + "</b></td></tr>")

            tw.write(
                " <tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Environment</b></td>")
            tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + TestEnv + "</b></td></tr>")

            tw.write(
                "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Test Start "
                "Time</b></td>")

            overall_startTime_str = overall_startTime.strftime("%d/%m/%Y %H:%M:%S")

            tw.write(
                " <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + str(
                    overall_startTime_str) + "</b></td></tr>")

            endTime = datetime.now()
            endTime_str = endTime.strftime("%d/%m/%Y %H:%M:%S")

            tw.write(
                "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Test End "
                "Time</b></td>")
            tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + str(endTime_str) + "</b></td></tr>")

            start_datetime = datetime.strptime(overall_startTime_str, "%d/%m/%Y %H:%M:%S")
            end_datetime = datetime.strptime(endTime_str, "%d/%m/%Y %H:%M:%S")

            tw.write(
                "<tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Total Test Ran "
                "Time</b></td>")
            tw.write(" <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + str(end_datetime - start_datetime) +
                     "</b></td></tr>")

            tw.write(
                " <tr><td bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2.75><b>Release</b></td>")
            tw.write(
                " <td><FONT COLOR=\"#153E7E\" FACE=\"Arial\" SIZE=2.75><b>" + release + "</b></td></tr></table></div>")

            tw.write(
                "<div class=\"topRight\"><h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5> Test Result Summary "
                ":</h4> <br><br>")

            tw.write("<table class =\"TestDetails\">")
            tw.write("<tr>")
            tw.write(
                "<td width=50% height=\"234px\" style=\"padding: 3px; border-right: double 2px lightgrey; "
                "vertical-align:text-top; \" align=\"center\">")

            tw.write(
                "<h3 class=\"heading\"> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=3.5> Test Modules Summary</h3>")

            tw.write("<table class=\"t06\" border=0 cellspacing=1 cellpadding=1 width=\"50\">")

            tw.write(
                "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Total Modules</b></td>")
            tw.write(
                "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">: ")
            tw.write(str(self.TotalModules) + "</b></td></tr>")
            tw.write(
                "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Total Pass</b></td>")
            tw.write(
                "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">: ")
            tw.write(str(self.numModulePass) + "</b></td></tr>")

            tw.write(
                "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Total Fail</b></td>")
            tw.write(
                "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">:  ")
            tw.write(str(self.numModuleFail) + "</b></td></tr></table>")

            tw.write(
                " <table class=\"t07 td02\" border=0 cellspacing=1 cellpadding=1><tr><td valign=\"bottom\"><b>" +
                str(successRateMod) + "%</b><div class=\"Passdiv\" style=\"height: " + str(passwidthMod) +
                "px; \"><span class=\"PercentageBarFont\"> " + str(self.numModulePass) + "</span></div></td></tr>")

            tw.write(
                "<tr><td style=\"height:1px;\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Pass</b></td></tr></table>")

            tw.write(
                "<table class=\"t08\" border=0 cellspacing=1 cellpadding=1><tr><td align = \"center\" "
                "valign=\"bottom\"><b>" + str(failRateMod) + "%</b><div class=\"Faildiv\" style=\"height: " + str(
                    failwidthMod) + "px; \"><span class=\"PercentageBarFont\">" + str(self.numModuleFail) +
                "</span></div></td></tr>")
            tw.write(
                "<tr><td style=\"height:1px;\" align=\"center\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Fail</b></td></tr></table>")

            tw.write("</td>")

            tw.write(
                "<td width=50% style=\"padding: 3px; vertical-align:text-top;\" align=\"center\"><h3 "
                "class=\"heading\"> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=3.5> Test Cases Summary</h3>")
            tw.write(
                "<table class=\"t06\" border=0 cellspacing=1 cellpadding=1 width=\"50\"><tr><td align=\"left\"><FONT "
                "COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b style=\"white-space:nowrap;\">Total Cases</b></td>")
            tw.write(
                "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">: " + str(NumTotal) + "</b></td></tr>")
            tw.write(
                "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Total Pass</b></td>")
            tw.write(
                "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">: " + str(self.FinalTestCasePass) + "</b></td></tr>")
            tw.write(
                "<tr><td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Total Fail</b></td>")
            tw.write(
                "<td align=\"left\"><FONT COLOR=\"#000066\" FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">: " + str(self.FinalTestCaseFail) + "</b></td></tr></table>")
            tw.write(
                "<table class=\"td02\" border=0 cellspacing=1 cellpadding=1><tr><td valign=\"bottom\" "
                "align=\"center\"><b>" + str(successRate) + "%</b>")
            tw.write(
                "<div class=\"Passdiv\" style=\"height: " + str(
                    passwidth) + "px; \"><span class=\"PercentageBarFont\">" + str(
                    self.FinalTestCasePass) + "</span></div></td></tr>")
            tw.write(
                "<tr><td valign=\"bottom\" style=\"height: 1px; \" align=\"center\"><FONT COLOR=\"#000066\" "
                "FACE=\"Arial\" SIZE=2.75><b style=\"white-space:nowrap;\">Pass</b></td></tr></table>")
            tw.write(
                "<table class=\"t08\" border=0 cellspacing=1 cellpadding=1><tr><td valign=\"bottom\" "
                "align=\"center\"><b>" + str(failRate) + "%</b>")
            tw.write(
                "<div class=\"Faildiv\" style=\"height: " + str(
                    failwidth) + "px; \"><span class=\"PercentageBarFont\">" + str(
                    self.FinalTestCaseFail) + "</span></div></td></tr>")
            tw.write(
                "<tr><td valign=\"bottom\" style=\"height: 1px; \" align=\"center\"><FONT COLOR=\"#000066\" "
                "FACE=\"Arial\" SIZE=2.75><b "
                "style=\"white-space:nowrap;\">Fail</b></td></tr></table></td></tr></table></div>")

            tw.write(
                "<div class=\"TopBottom\"> <h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5>  Detailed Test Report "
                ":</h4>")

            tw.write(
                "<table class =\"DetailedReport\" border=1 cellspacing=1 cellpadding=1 width=100% align='center'>")  # <tr>")

            tw.write(
                "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                "SIZE=2><b>Browser</b></td>")

            tw.write(
                "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Execution "
                "Time</b></td>")
            tw.write(
                "<td align=\"center\" bgcolor=\"#153E7E\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                "SIZE=2><b>Status</b></td>")

            tw.write(" </tr>")

            tw.write(self.html_body_string)

            tw.write("</table></div></body></html>")

            tw.close()

    def Status_Color(self, status1):
        try:
            status = status1.lower()

            if status == "pass":
                self.StatusDetbgcolor = "#008000"
            elif status == "fail":
                self.StatusDetbgcolor = "#f71919"
            elif status == "not executed":
                self.StatusDetbgcolor = "#FBD105"
            elif status == "done":
                self.StatusDetbgcolor = "#BCE954"
        except Exception as e:
            print("'Status_Color' has this error -> \n" + self.common_methods.error_message(e))
        return self.StatusDetbgcolor

    # This is workig code, create a file with browser and module details
    def Report_BrowserSummary_working(self, browsername, browser_startTime_str):
        try:
            rstatus = "Pass"
            self.folders()

            #Report_Browsers_Filename = str(self.TestBrowsersFolder) + "\\" + "Browsers.html"
            Report_Browsers_Filename = os.path.join(str(self.TestBrowsersFolder), "Browsers.html")

            if self.numModuleFail > 0:
                BrowserStatus = "Fail"
                self.StatusDetbgcolor = self.Status_Color("fail")
            else:
                BrowserStatus = "Pass"
                self.StatusDetbgcolor = self.Status_Color("pass")

            if not os.path.exists(Report_Browsers_Filename):
                # make a new file if not
                tw = open(Report_Browsers_Filename, 'w')

                tw.write(self.new_html_styles_part1)

                tw.write("<ul>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestReport_highReport.html\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                tw.write(
                    "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\AboutTheProject\\AboutTheProject.html\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\About\\About.html\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                tw.write("</ul>")

                tw.write(self.new_html_styles_part3)

                tw.write("<div class=\"content-header\">")
                tw.write("<h1> Browsers </h1>")
                tw.write("<p>Test Automation Report - Summary of Test Suites</p>")
                tw.write("</div>")

                tw.write(
                    "<div class=\"TopBottom\"> <h4> <FONT COLOR=\"660000\" FACE=\"Arial\" SIZE=4.5>  Detailed Test "
                    "Report:</h4>")
                tw.write(
                    "<table class =\"TestDetails\" border=1 cellspacing=1 cellpadding=1 width=100% align='center'><tr>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                    "SIZE=2><b>Browser</b></td>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Total "
                    "Module(s) </b></td>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Module("
                    "s) Passed "
                    "</b></td>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Module("
                    "s) Failed "
                    "</b></td>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" SIZE=2><b>Execution Time "
                    "</b></td>")
                tw.write(
                    "<td align=\"center\" bgcolor=\"#0274BD\"><FONT COLOR=\"#E0E0E0\" FACE=\"Arial\" "
                    "SIZE=2><b>Status</b></td>")

                tw.write(" </tr>")

                # print("htmlBody -> " + htmlBody)
                # tw.write(htmlBodyBrowser)

                tw.write(
                    "<tr><td width=20% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b><a href=..\\TestModules\\" + \
                    self.testmodulefilepath + ">" + browsername + "</a></b></td>")

                # print("str(self.numModulePass+self.numModuleFail) -> " + str(self.numModulePass + self.numModuleFail))
                # print("self.numModulePass -> " + str(self.numModulePass))
                # print("self.numModuleFail) -> " + str(self.numModuleFail))

                tw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                         + str(self.numModulePass + self.numModuleFail) + "</b></td>")

                tw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                         + str(self.numModulePass) + "</b></td>")

                tw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                         + str(self.numModuleFail) + "</b></td>")

                browser_endTime = datetime.datetime.now()
                browser_endTime_str = browser_endTime.strftime("%d/%m/%Y %H:%M:%S")

                browser_start_datetime = datetime.datetime.strptime(browser_startTime_str, "%d/%m/%Y %H:%M:%S")
                browser_end_datetime = datetime.datetime.strptime(browser_endTime_str, "%d/%m/%Y %H:%M:%S")

                tw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                         + str(browser_end_datetime - browser_start_datetime) + "</b></td>")

                tw.write("<td width=15% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT COLOR=\"#FFFFFF\" " \
                                                                                             "FACE=\"Arial\" SIZE=2><b>" + \
                         BrowserStatus + "</b></td></tr> ")

                tw.close()

            else:
                rbw = open(Report_Browsers_Filename, 'a')
                # print("htmlBody rbw already exist -> " + htmlBody)

                # rbw.write(htmlBodyBrowser)

                rbw.write(
                    "<tr><td width=20% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b><a href=" + \
                    self.Report_TestModule_Filename + ">" + browsername + "</a></b></td>")

                # print("str(self.numModulePass+self.numModuleFail) -> " + str(self.numModulePass + self.numModuleFail))
                # print("self.numModulePass -> " + str(self.numModulePass))
                # print("self.numModuleFail) -> " + str(self.numModuleFail))

                rbw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                          + str(self.numModulePass + self.numModuleFail) + "</b></td>")

                rbw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                          + str(self.numModulePass) + "</b></td>")

                rbw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                          + str(self.numModuleFail) + "</b></td>")

                browser_endTime = datetime.datetime.now()
                browser_endTime_str = browser_endTime.strftime("%d/%m/%Y %H:%M:%S")

                browser_start_datetime = datetime.datetime.strptime(browser_startTime_str, "%d/%m/%Y %H:%M:%S")
                browser_end_datetime = datetime.datetime.strptime(browser_endTime_str, "%d/%m/%Y %H:%M:%S")

                rbw.write("<td width=15% align=\"center\"><FONT COLOR=\"#000000\" FACE=\"Arial\" SIZE=2><b>" \
                          + str(browser_end_datetime - browser_start_datetime) + "</b></td>")

                rbw.write(
                    "<td width=15% align=\"center\" bgcolor=" + self.StatusDetbgcolor + "><FONT COLOR=\"#FFFFFF\" " \
                                                                                        "FACE=\"Arial\" SIZE=2><b>" + \
                    BrowserStatus + "</b></td></tr> ")

                rbw.close()
        except Exception as e:
            print(
                "'Report_BrowserSummary_working' Action Exception Message -> \n" + self.common_methods.error_message(e))

    def about_the_project(self, projectName):
        try:
            #abouttheproject_Filename = str(self.about_the_project_Folder) + "\\" + "AboutTheProject.html"
            abouttheproject_Filename = os.path.join(str(self.about_the_project_Folder), "AboutTheProject.html")

            new_html_styles_part3 = "</div>" + \
                                    "</section>" + \
                                    "<section id=\"content-div\">" + \
                                    "<div id=\"header\">" + \
                                    "<div class=\"header-nav\">" + \
                                    "<div class=\"menu-button\">Test Automation Report - " + projectName + "</div>" + \
                                    "<div class=\"nav\">" + \
                                    "<ul>" + \
                                    "<li class=\"nav-profile\">" + \
                                    "<div class=\"nav-profile-name\">" + \
                                    os.getlogin() + "<i class=\"fa fa-caret-down\"></i>" + \
                                    "</div>" + \
                                    "</li>" + \
                                    "</ul>" + \
                                    "</div>" + \
                                    "</div>" + \
                                    "</div>" + \
                                    "<div class=\"content\" id=\"dashboard\" style=\"display: block;\">"

            if not os.path.exists(abouttheproject_Filename):
                # make a new file if not
                tw = open(abouttheproject_Filename, 'w')

                # tw.write(header)

                tw.write("<html><head> ")
                tw.write("<title>Selenium Python Automation</title>")
                tw.write("<link rel=\"stylesheet\" href=\"../css/style.css\">")
                tw.write("<script type=\"text/javascript\">")

                tw.write("</script>")
                tw.write("</head>")
                tw.write("<body>")
                tw.write("<section id=\"sidebar\">")
                tw.write("<div class=\"white-label\">")
                tw.write("</div>")
                tw.write("<div class=\"container\" id=\"sidebar-nav\">")

                tw.write("<ul>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestReport_highReport.html\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")
                tw.write(
                    "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\About\\About.html\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                tw.write("</ul>")

                tw.write(new_html_styles_part3)

                tw.write("<div class=\"content-header\">")
                tw.write("<h1>About the Project</h1>")
                tw.write("<p>Test Automation Report - Summary of the project</p>")
                tw.write("</div>")
                tw.write("<p> ")
                # tw.write("CyberGrants is a cloud-based philanthropy management software platform that provides "
                #        "organizations with a scalable and flexible solution for managing their corporate social "
                 #        "responsibility (CSR) and philanthropic initiatives. ")


                #tw.write("<br><br>As a SaaS offering, CyberGrants "
                #         "provides a comprehensive suite of tools and features that enable organizations to "
                #        "streamline their grantmaking processes, manage employee giving and volunteering programs, "
                #        "and track and report on the impact of their philanthropic activities, all without the need "
                #        "for on-premises software installations or maintenance. ")
                #tw.write("<br><br>By leveraging the power of the "
                #        "cloud, CyberGrants enables organizations to operate more efficiently, with increased "
                #        "collaboration and transparency across teams, while also reducing costs and increasing "
                #        "overall agility.")
                tw.write("Clinical Trial Knowledge Management System (CTKMS) or Knowledge Management System (KMS) is a single, central maintained repository with defined data owners that holds key country clinical trial information.")
                tw.write("<br><br>The system ensures data remains up-to-date and compliant with country-specific legislation")
                tw.write("<br><br>The project features two distinct portals:")
                tw.write("<br><br>Admin Portal: Used for managing configurations and overseeing data control.")
                tw.write("<br><br>End-User Portal: Designed for easy access to reference clinical trial information.")
                tw.write("<br><br>KMS is built on a serverless microservices architecture and is fully deployed on AWS, ensuring scalability, cost-efficiency, and ease of maintenance.")
                tw.write("<br><br><br>KMS can host multiple workspaces in it.  Repository of Clinical trials Knowledge (ROCK) is one such workspace.")
                tw.write("</p>")
                tw.close()
        except Exception as e:
            print()

    def about_report(self, projectName):
        try:
            #about_Filename = str(self.aboutFolder) + "\\" + "About.html"
            about_Filename = os.path.join(str(self.aboutFolder), "About.html")

            new_html_styles_part3 = "</div>" + \
                                    "</section>" + \
                                    "<section id=\"content-div\">" + \
                                    "<div id=\"header\">" + \
                                    "<div class=\"header-nav\">" + \
                                    "<div class=\"menu-button\">Test Automation Report - " + projectName + "</div>" + \
                                    "<div class=\"nav\">" + \
                                    "<ul>" + \
                                    "<li class=\"nav-profile\">" + \
                                    "<div class=\"nav-profile-name\">" + \
                                    os.getlogin() + "<i class=\"fa fa-caret-down\"></i>" + \
                                    "</div>" + \
                                    "</li>" + \
                                    "</ul>" + \
                                    "</div>" + \
                                    "</div>" + \
                                    "</div>" + \
                                    "<div class=\"content\" id=\"dashboard\" style=\"display: block;\">"

            if not os.path.exists(about_Filename):
                # make a new file if not
                tw = open(about_Filename, 'w')

                # tw.write(header)

                tw.write("<html><head> ")
                tw.write("<title>Selenium Python Automation</title>")
                tw.write("<link rel=\"stylesheet\" href=\"../css/style.css\">")
                tw.write("<script type=\"text/javascript\">")

                tw.write("</script>")
                tw.write("</head>")
                tw.write("<body>")
                tw.write("<section id=\"sidebar\">")
                tw.write("<div class=\"white-label\">")
                tw.write("</div>")
                tw.write("<div class=\"container\" id=\"sidebar-nav\">")

                tw.write("<ul>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestReport_highReport.html\" class=\"toggle-btn\" data-content=\"dashboard\">Dashboard</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\Browsers\\Browsers.html\" class=\"toggle-btn\" data-content=\"test-suites\">Browsers</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestModules\\TestModules.html\" class=\"toggle-btn\" data-content=\"test-suites\">Test Suites</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\TestCases\\TCModuleName.html\" class=\"toggle-btn\" data-content=\"test-cases\">Test Cases</a></li>")
                tw.write(
                    "<li class=\"leftnav\"><a href=\"..\\AboutTheProject\\AboutTheProject.html\" class=\"toggle-btn\" data-content=\"about-the-project\">About the project</a></li>")
                tw.write(
                    "<li class=\"leftnav active\"><a href=\"#\" class=\"toggle-btn\" data-content=\"about\">About Test Automation Tool</a></li>")
                tw.write("</ul>")

                tw.write(new_html_styles_part3)

                tw.write("<div class=\"content-header\">")
                tw.write("<h1>About</h1>")
                tw.write("<p>Test Automation Report - Summary of the project</p>")
                tw.write("</div>")
                tw.write("<p> ")
                tw.write("Selenium is a widely-used open-source tool for automating web browsers.")
                tw.write("<br>It provides "
                         "a powerful and flexible way to automate testing for user interactions on the web applications.")
                tw.write("<br>Python is a widely-used programming language know for its clean and readable syntax, making it easier to write and maintain automation scripts.")
                tw.write(
                    "Python has a large collection of libraries and modules that can be used for various automation tasks.")
                tw.write("<br><br>Version of Test Automation Framework:")
                tw.write("Selenium Python Automation Version 1.5 - 23rd May 2023")
                tw.write("</p>")
                tw.close()
        except Exception as e:
            print()

    def writeCSS(self):
        try:
            # Open the file for writing
            cssName = "style.css"
            #css_filename = str(self.cssFolder) + "\\" + cssName
            css_filename = os.path.join(str(self.cssFolder), cssName)

            with open(css_filename, 'w') as f:
                # Write the data to the file
                f.write("""
                        html,
                        body {
                            font-family: -apple-system, sans-serif;
                            height: 100%;
                        }
                        body {
                            background: #FFFFFF;
                            height: 100%;
                        }
                        img {
                            max-width: 100%;
                        }
                        ul {
                            list-style: none;
                            margin: 0;
                            padding: 0;
                        }
                        a {
                            text-decoration: none;
                        }
                        #header {
                            float: left;
                            width: 100%;
                            background: #ffffff;
                            position: relative;
                        }
                        .white-label {
                            float: left;
                            background: #33373B;
                            max-width: 210px;
                            padding: 10px;
                            min-height: 44px;
                            background: #279BE4;
                            width: 100%;
                            max-height: 44px;
                        }
                        .white-label img {
                            max-height: 43px;
                        }
                        .header-nav {
                            min-height: 64px;
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                            background: #279BE4;
                        }
                        .menu-button {
                            float: left;
                            font-size: 29px;
                            color: #fff;
                            padding: 12px 19px;
                        }
                        .nav ul {
                            height: 64px;
                            float: right;
                        }
                        .nav ul li {
                            float: left;
                            position: relative;
                            padding: 11px;
                        }
                        .nav > ul > li:first-child {
                            border-left: none;
                        }
                        .nav ul li a {
                            color: #fff;
                            padding: 1px;
                            float: left;
                        }
                        .nav ul li i {
                            color: #fff;
                        }
                        .nav ul li:hover {
                            background: #01A9F0;
                            color: #fff;
                        }
                        .user-profile {
                            float: right;
                        }
                        .user-profile > div {
                            float: left;
                            padding: 20px 8px;
                            position: relative;
                        }
                        .user-profile i {
                            font-size: 1.2em;
                            color: #5F6F86;
                        }
                        .user-profile i:hover {
                            color: #397AC5;
                        }
                        .font-icon i:after {
                            position: absolute;
                            content: "3";
                            background: #E74C3C;
                            color: #fff;
                            font-size: 12px;
                            border-radius: 50%;
                            width: 10px;
                            height: 10px;
                            padding: 3px 4px 4px 3px;
                            text-align: center;
                            top: 12px;
                            right: 11px;
                        }
                        .font-icon {
                            padding: 8px 10px;
                        }
                        .font-icon i {
                            font-size: 24px;
                        }
                        .nav-mail .font-icon i:after {
                            background: #2ECC71;
                        }
                        div.user-image {
                            padding: 9px 5px;
                            margin: 0 5px;
                            border-left: 1px solid #ccc;
                            border-right: 1px solid #ccc;
                        }
                        .nav-profile {
                            background: #0274BD;
                        }
                        .nav-profile-image img {
                            width: 39px;
                            height: 41px;
                            border-radius: 50%;
                            float: left;
                        }
                        .nav-profile-name {
                            float: right;
                            margin: 11px 7px 8px 14px;
                            color: #fff;
                        }
                        .nav-profile-name i {
                            padding: 0 0 0 11px;
                        }
                        .nav-chat i:after {
                            display: none;
                        }
                        #sidebar {
                            overflow: hidden;
                            width: 210px;
                            height: 100%;
                            float: left;
                            background: #2A2D33;
                        }
                        #sidebar-nav {
                            width: 106%;
                            height: calc(100% - 95px);
                            padding: 0;
                            background: #2A2D33;
                            border-right: 1px solid #E0E0E0;
                            overflow-y: scroll;
                        }
                        #sidebar-nav h2 {
                            color: #60636B;
                            float: left;
                            width: 100%;
                            font-size: 0.8em;
                            font-family: -apple-system, sans-serif;
                            font-weight: 600;
                            text-transform: uppercase;
                            padding: 3px 0 2px 20px;
                            border-top: 1px solid #4D4C4C;
                            box-sizing: border-box;
                            margin: 10px 0;
                        }
                        #sidebar-nav ul {
                        }
                        #sidebar-nav ul li {
                        }
                        #sidebar-nav ul li a {
                            color: #C2C2C2;
                            font-size: 0.95em;
                            padding: 15px 20px;
                            float: left;
                            width: 100%;
                            font-weight: 600;
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                        }
                        #sidebar-nav ul li:hover a,
                        #sidebar-nav ul li:hover a i,
                        #sidebar-nav li.active a,
                        #sidebar-nav li.active a i {
                            color: #333;
                        }
                        #sidebar-nav ul li:hover a {
                            background: #fff;
                            color: #333;
                        }
                        #sidebar-nav ul li.active a {
                            background: #fff;
                            color: #333;
                        }
                        #sidebar-nav ul li.active a i {
                            background: #fff;
                        }
                        #sidebar-nav i {
                            padding-right: 8px;
                            font-size: 1.3em;
                            color: #60636B;
                            width: 25px;
                            text-align: center;
                        }
                        #content-div {
                            float: left;
                            width: calc(100% - 210px);
                            height: 100%;
                            word-wrap: break-word;
                            background: #FFFFFF;
                            font-family: -apple-system, sans-serif;
                        }
                        ::-webkit-scrollbar {
                            width: 12px;
                        }
                        ::-webkit-scrollbar-track {
                            -webkit-box-shadow: inset 0 0 6px rgba(0, 0, 0, 0.3);
                            border-radius: 10px;
                        }
                        ::-webkit-scrollbar-thumb {
                            border-radius: 10px;
                            -webkit-box-shadow: inset 0 0 6px rgba(0, 0, 0, 0.5);
                        }
                        .content-div {
                            float: left;
                            background: #E9EEF4;
                            width: 100%;
                            height: calc(100% - 64px);
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                        }
                        .content-header {
                            background: #fff;
                            float: left;
                            width: 100%;
                            margin-bottom: 15px;
                            padding: 15px;
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                            border-bottom: 1px solid #ccc;
                        }
                        .content-header h1 {
                            margin: 0;
                            font-weight: normal;
                            padding-bottom: 5px;
                        }
                        .content-header p {
                            margin: 0;
                            padding-left: 2px;
                        }
                        .widget-box {
                            background: #fff;
                            border: 1px solid #E0E0E0;
                            float: left;
                            width: 100%;
                            margin: 0 0 15px 15px;
                        }
                        .widget-header {
                            background: #279BE4;
                        }
                        .widget-header h2 {
                            font-size: 15px;
                            font-weight: normal;
                            margin: 0;
                            padding: 11px 15px;
                            color: #F9F9F9;
                            display: inline-block;
                        }
                        .sample-widget {
                            max-width: 47%;
                        }
                        .widget-box .fa-cog {
                            float: right;
                            color: #fff;
                            margin: 11px 11px 0 0;
                            font-size: 20px;
                        }

                        body {
                            font-family: -apple-system, sans-serif;
                            height: 100%;
                        }
                        body {
                            background: #FFFFFF;
                            height: 100%;
                        }
                        img {
                            max-width: 100%;
                        }
                        ul {
                            list-style: none;
                            margin: 0;
                            padding: 0;
                        }
                        a {
                            text-decoration: none;
                        }

                        #header {
                            float: left;
                            width: 100%;
                            background: #ffffff;
                            position: relative;
                        }
                        .white-label {
                            float: left;
                            background: #33373B;
                            max-width: 210px;
                            padding: 10px;
                            min-height: 44px;
                            background: #279BE4;
                            width: 100%;
                            max-height: 44px;
                        }
                        .white-label img {
                            max-height: 43px;
                        }
                        .header-nav {
                            min-height: 64px;
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                            background: #279BE4;
                        }
                        .menu-button {
                            float: left;
                            font-size: 29px;
                            color: #fff;
                            padding: 12px 19px;
                        }
                        .nav ul {
                            height: 64px;
                            float: right;
                        }
                        .nav ul li {
                            float: left;
                            position: relative;
                            padding: 11px;
                        }
                        .nav > ul > li:first-child {
                            border-left: none;
                        }
                        .nav ul li a {
                            color: #fff;
                            padding: 1px;
                            float: left;
                        }
                        .nav ul li i {
                            color: #fff;
                        }
                        .nav ul li:hover {
                            background: #01A9F0;
                            color: #fff;
                        }
                        .user-profile {
                            float: right;
                        }
                        .user-profile > div {
                            float: left;
                            padding: 20px 8px;
                            position: relative;
                        }
                        .user-profile i {
                            font-size: 1.2em;
                            color: #5F6F86;
                        }
                        .user-profile i:hover {
                            color: #397AC5;
                        }
                        .font-icon i:after {
                            position: absolute;
                            content: "3";
                            background: #E74C3C;
                            color: #fff;
                            font-size: 12px;
                            border-radius: 50%;
                            width: 10px;
                            height: 10px;
                            padding: 3px 4px 4px 3px;
                            text-align: center;
                            top: 12px;
                            right: 11px;
                        }
                        .font-icon {
                            padding: 8px 10px;
                        }
                        .font-icon i {
                            font-size: 24px;
                        }
                        .nav-mail .font-icon i:after {
                            background: #2ECC71;
                        }
                        div.user-image {
                            padding: 9px 5px;
                            margin: 0 5px;
                            border-left: 1px solid #ccc;
                            border-right: 1px solid #ccc;
                        }
                        .nav-profile {
                            background: #0274BD;
                        }
                        .nav-profile-image img {
                            width: 39px;
                            height: 41px;
                            border-radius: 50%;
                            float: left;
                        }
                        .nav-profile-name {
                            float: right;
                            margin: 11px 7px 8px 14px;
                            color: #fff;
                        }
                        .nav-profile-name i {
                            padding: 0 0 0 11px;
                        }
                        .nav-chat i:after {
                            display: none;
                        }

                        #sidebar {
                            overflow: hidden;
                            width: 210px;
                            height: 100%;
                            float: left;
                            background: #2A2D33;
                        }
                        #sidebar-nav {
                            width: 106%;
                            height: calc(100% - 95px);
                            padding: 0;
                            background: #2A2D33;
                            border-right: 1px solid #E0E0E0;
                            overflow-y: scroll;
                        }
                        #sidebar-nav h2 {
                            color: #60636B;
                            float: left;
                            width: 100%;
                            font-size: 0.8em;
                            font-family: -apple-system, sans-serif;
                            font-weight: 600;
                            text-transform: uppercase;
                            padding: 3px 0 2px 20px;
                            border-top: 1px solid #4D4C4C;
                            box-sizing: border-box;
                            margin: 10px 0;
                        }
                        #sidebar-nav ul {
                        }
                        #sidebar-nav ul li {
                        }
                        #sidebar-nav ul li a {
                            color: #C2C2C2;
                            font-size: 0.95em;
                            padding: 15px 20px;
                            float: left;
                            width: 100%;
                            font-weight: 600;
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                        }
                        #sidebar-nav ul li:hover a,
                        #sidebar-nav ul li:hover a i,
                        #sidebar-nav li.active a,
                        #sidebar-nav li.active a i {
                            color: #333;
                        }
                        #sidebar-nav ul li:hover a {
                            background: #fff;
                            color: #333;
                        }
                        #sidebar-nav ul li.active a {
                            background: #fff;
                            color: #333;
                        }
                        #sidebar-nav ul li.active a i {
                            background: #fff;
                        }
                        #sidebar-nav i {
                            padding-right: 8px;
                            font-size: 1.3em;
                            color: #60636B;
                            width: 25px;
                            text-align: center;
                        }

                        .content {
                            float: left;
                            background: #E9EEF4;
                            width: 100%;
                            /* this is for grey background, removing this applies to the entire page,even the left nav extends */
                            /* height: calc(100% - 64px); */
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                        }
                        .content-header {
                            background: #fff;
                            float: left;
                            width: 100%;
                            margin-bottom: 15px;
                            padding: 15px;
                            -webkit-box-sizing: border-box;
                            -moz-box-sizing: border-box;
                            box-sizing: border-box;
                            border-bottom: 1px solid #ccc;
                        }
                        .content-header h1 {
                            margin: 0;
                            font-weight: normal;
                            padding-bottom: 5px;
                        }
                        .content-header p {
                            margin: 0;
                            padding-left: 2px;
                        }
                        .widget-box {
                            background: #fff;
                            border: 1px solid #E0E0E0;
                            float: left;
                            width: 100%;
                            margin: 0 0 15px 15px;
                        }
                        .widget-header {
                            background: #279BE4;
                        }
                        .widget-header h2 {
                            font-size: 15px;
                            font-weight: normal;
                            margin: 0;
                            padding: 11px 15px;
                            color: #F9F9F9;
                            display: inline-block;
                        }
                        .sample-widget {
                            max-width: 47%;
                        }
                        .widget-box .fa-cog {
                            float: right;
                            color: #fff;
                            margin: 11px 11px 0 0;
                            font-size: 20px;
                        }

                        /* --- this is old report Test Details section CSS- works
                        .TestDetails td {
                            border: 1px solid #ccc;
                            text-align: left;
                            padding: 7px;
                        }

                        .TestDetails th {border: 1px solid #ccc;text-align: left;padding: 7px;}

                        .TestDetails td:first-child {width: 150px;}

                        .TestDetails th {background: lightblue;border-color: white;}

                        .TestDetails tr:nth-child(even) {background-color: #f2f2f2}body {padding: 1rem;}
                        */

                         .TestDetails {
                            border-collapse: collapse;
                            margin: 2px 0;
                            font-size: 0.9em;
                            font-family: -apple-system, sans-serif;
                            min-width: 400px;
                            box-shadow: 0 0 20px rgba(0, 0, 0, 0.15);
                        }
                        .TestDetails thead tr {
                            background-color: #009879;
                            color: #ffffff;
                            text-align: left;
                        }
                        .TestDetails th,
                        .TestDetails td {
                            padding: 10px 15px;
                        }
                        .TestDetails tbody tr {
                            border-bottom: 1px solid #dddddd;
                        }

                        .TestDetails tbody tr:nth-of-type(even) {
                            background-color: #f3f3f3;
                        }

                        .TestDetails tbody tr:last-of-type {
                            border-bottom: 1px solid #009879;
                        }
                        .TestDetails tbody tr.active-row {
                            font-weight: bold;
                            color: #009879;
                        }

                        /* Same style is been applied to the table, it is not applying if given as .TestDetails, .t06 */

                        .t06 {
                            border-collapse: collapse;
                            margin: 2px 0;
                            font-size: 0.9em;
                            font-family: -apple-system, sans-serif;
                            box-shadow: 0 0 20px rgba(0, 0, 0, 0.15);
                        }
                        .t06 thead tr {
                            background-color: #009879;
                            color: #ffffff;
                            text-align: left;
                        }
                        .t06 th,
                        .t06 td {
                            padding: 10px 15px;
                        }
                        .t06 tbody tr {
                            border-bottom: 1px solid #dddddd;
                        }

                        .t06 tbody tr:nth-of-type(even) {
                            background-color: #f3f3f3;
                        }

                        .t06 tbody tr:last-of-type {
                            border-bottom: 1px solid #009879;
                        }
                        .t06 tbody tr.active-row {
                            font-weight: bold;
                            color: #009879;
                        }

                        .ModulePerTable {
                          position: relative;
                        }

                        .t07 {
                          bottom: 0;
                        }

                        .t06{
                          margin-left: auto;
                          margin-right: auto;
                          }

                        .t07{
                          margin-left: auto;
                          margin-right: auto;
                          }

                        .t08{
                          margin-left: auto;
                          margin-right: auto;
                          }

                        select {
                          padding: 10px;
                          font-size: 16px;
                          border: 1px solid #ccc;
                          border-radius: 5px;
                          background-color: #f1f1f1;
                          margin-right: 10px;
                          margin-left: 10px;
                          width: 200px;
                        }

                        button {
                          background-color: #4CAF50;
                          color: white;
                          padding: 10px 20px;
                          border: none;
                          border-radius: 5px;
                          cursor: pointer;
                        }

                        button:hover {
                          background-color: #3e8e41;
                        }

                        .search-results-message {
                          margin-top: 20px;
                          padding: 10px;
                          border: 1px solid #ccc;
                          border-radius: 5px;
                          background-color: #f1f1f1;
                          text-align: center;
                          display: none;
                        }

                        h4{
                             margin-left: 10px;
                        }
                        
                        .tabbed-widget {
                            /* border: 1px solid #ccc; */
                            padding: 10px;
						}
						
						.tabs {
						  list-style: none;
						  margin: 0;
						  padding: 0;
						  border-bottom: 1px solid #ccc;
						  display: flex;
						}
						
						.tabs li {
						  margin-right: 10px;
						}
						.tabs a {
						  display: block;
						  padding: 10px;
						  background-color: #eee;
						  border: 1px solid #ccc;
						  border-bottom: none;
						}
						.tabs a:hover {
						  background-color: #ddd;
						}
						.tabs a.active {
						  background-color: #fff;
						  border-bottom: 1px solid #fff;
						}
						.tab-content {
						  display: none;
						}
						.tab-content.active {
						  display: block;
						}
						/* Style the tooltip text */
						.tooltip {
						  position: relative;
						  display: inline-block;
						}

						/* Style the tooltip content */
						.tooltip .tooltiptext {
						  visibility: hidden;
						  width: 150px;
						  background-color: #fff;
						  color:#555;
						  text-align: center;
						  padding: 5px 0;
						  border-radius: 6px;
						  
						  /* Position the tooltip content */
						  position: absolute;
						  z-index: 1;
						  bottom: 125%;
						  left: 50%;
						  margin-left: -60px;
						}

						/* Show the tooltip text when you hover over the tooltip container */
						.tooltip:hover .tooltiptext {
						  visibility: visible;
						}
						.tabbed-widget-container{
							/* height:239px; */
							overflow-y:scroll;
						}
						
						.hidden{
							display:none;
						}
                        """)
        except Exception as e:
            print(self.common_methods.error_message(e))
